# Databricks notebook source
# ============================================================
# CELL 1 | INSTALLATION
# Run once when setting up
# ============================================================

import subprocess, sys

PACKAGES = [
    "requests", "beautifulsoup4", "pandas",
    "plotly", "folium", "lxml",
    "gnews", "fake-useragent", "python-dotenv",
]

print("📦 Installing packages...")
for pkg in PACKAGES:
    r = subprocess.run(
        [sys.executable, "-m", "pip", "install", pkg, "-q"],
        capture_output=True, text=True
    )
    print(f"  {'✅' if r.returncode == 0 else '❌'} {pkg}")

print("\n✅ Installation complete!")

# COMMAND ----------

# ============================================================
# CELL 2 | IMPORTS
# Run every session
# ============================================================

# Standard library
import os, json, time, sqlite3, logging
from pathlib import Path
from datetime import datetime, timedelta
from contextlib import contextmanager

# Third party
import requests
import pandas as pd
from bs4 import BeautifulSoup

# Optional imports with graceful fallback
try:
    from gnews import GNews
    GNEWS_OK = True
except ImportError:
    GNEWS_OK = False
    print("⚠️  gnews not available — RSS fallback will be used")

try:
    from fake_useragent import UserAgent
    _ua = UserAgent()
    def random_ua():
        try:    return _ua.random
        except: return "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"
except ImportError:
    def random_ua():
        return "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"

print("✅ All imports ready")

# COMMAND ----------

# ============================================================
# CELL 3 | CONFIGURATION
# ⚠️  THIS IS THE ONLY CELL YOU NEED TO EDIT
# ============================================================

# ----------------------------------------------------------
# API KEY
# Get free key → https://openweathermap.org/api
OPENWEATHER_API_KEY = os.getenv("OPENWEATHER_API_KEY", "")

# ----------------------------------------------------------
# CITIES
# Add / remove cities here as needed
# ----------------------------------------------------------
INDIA_CITIES = {
    # ── Metro ──────────────────────────────────────────────
    "Mumbai":      {"lat": 19.0760, "lon": 72.8777, "state": "Maharashtra",      "tier": "Metro"},
    "Delhi":       {"lat": 28.6139, "lon": 77.2090, "state": "Delhi",            "tier": "Metro"},
    "Bangalore":   {"lat": 12.9716, "lon": 77.5946, "state": "Karnataka",        "tier": "Metro"},
    "Chennai":     {"lat": 13.0827, "lon": 80.2707, "state": "Tamil Nadu",       "tier": "Metro"},
    "Kolkata":     {"lat": 22.5726, "lon": 88.3639, "state": "West Bengal",      "tier": "Metro"},
    "Hyderabad":   {"lat": 17.3850, "lon": 78.4867, "state": "Telangana",        "tier": "Metro"},
    # ── Tier 1 ─────────────────────────────────────────────
    "Pune":        {"lat": 18.5204, "lon": 73.8567, "state": "Maharashtra",      "tier": "Tier-1"},
    "Ahmedabad":   {"lat": 23.0225, "lon": 72.5714, "state": "Gujarat",          "tier": "Tier-1"},
    "Jaipur":      {"lat": 26.9124, "lon": 75.7873, "state": "Rajasthan",        "tier": "Tier-1"},
    "Lucknow":     {"lat": 26.8467, "lon": 80.9462, "state": "Uttar Pradesh",    "tier": "Tier-1"},
    # ── Tier 2 ─────────────────────────────────────────────
    "Chandigarh":  {"lat": 30.7333, "lon": 76.7794, "state": "Punjab",           "tier": "Tier-2"},
    "Bhopal":      {"lat": 23.2599, "lon": 77.4126, "state": "Madhya Pradesh",   "tier": "Tier-2"},
    "Patna":       {"lat": 25.5941, "lon": 85.1376, "state": "Bihar",            "tier": "Tier-2"},
    "Kochi":       {"lat":  9.9312, "lon": 76.2673, "state": "Kerala",           "tier": "Tier-2"},
    "Guwahati":    {"lat": 26.1445, "lon": 91.7362, "state": "Assam",            "tier": "Tier-2"},
    "Bhubaneswar": {"lat": 20.2961, "lon": 85.8245, "state": "Odisha",           "tier": "Tier-2"},
    "Srinagar":    {"lat": 34.0837, "lon": 74.7973, "state": "J&K",              "tier": "Tier-2"},
    "Shimla":      {"lat": 31.1048, "lon": 77.1734, "state": "Himachal Pradesh", "tier": "Tier-2"},
    "Dehradun":    {"lat": 30.3165, "lon": 78.0322, "state": "Uttarakhand",      "tier": "Tier-2"},
    "Raipur":      {"lat": 21.2514, "lon": 81.6296, "state": "Chhattisgarh",     "tier": "Tier-2"},
}

# ----------------------------------------------------------
# STRIKE KEYWORDS
# Add more keywords if needed
# ----------------------------------------------------------
STRIKE_KEYWORDS = [
    "transport strike", "bus strike", "truck strike",
    "auto strike",      "taxi strike", "rail strike",
    "railway strike",   "lorry strike", "driver strike",
    "bandh",            "hartal", "chakka jam",
    "bharat bandh",     "road blockade",
]

# ----------------------------------------------------------
# STORAGE PATHS
# ----------------------------------------------------------
BASE_DIR = Path("india_monitor_data")
DB_PATH  = BASE_DIR / "database" / "india_monitor.db"
CSV_DIR  = BASE_DIR / "csv_exports"
JSON_DIR = BASE_DIR / "json_exports"
LOG_DIR  = BASE_DIR / "logs"

# Create all folders
for d in [DB_PATH.parent, CSV_DIR, JSON_DIR, LOG_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ----------------------------------------------------------
# ALERT THRESHOLDS
# ----------------------------------------------------------
ALERT_THRESHOLDS = {
    "rain_mm_per_hr" : 20,    # Heavy rain
    "temp_extreme_c" : 42,    # Extreme heat
    "wind_ms"        : 15,    # Strong wind
    "visibility_km"  : 1,     # Low visibility
}

# ----------------------------------------------------------
# VALIDATION
# ----------------------------------------------------------
print("=" * 50)
print("⚙️  CONFIGURATION SUMMARY")
print("=" * 50)
print(f"  API Key   : {'✅ Set' if OPENWEATHER_API_KEY != 'your_api_key_here' else '⚠️  Not set (historical still works)'}")
print(f"  Cities    : {len(INDIA_CITIES)}")
print(f"  Keywords  : {len(STRIKE_KEYWORDS)}")
print(f"  Database  : {DB_PATH}")
print("=" * 50)

RAG_DIR        = BASE_DIR / "rag"
DOCS_DIR       = RAG_DIR  / "documents"
VECTOR_DIR     = RAG_DIR  / "vector_store"
PROCESSED_DIR  = RAG_DIR  / "processed"

for d in [DOCS_DIR, VECTOR_DIR, PROCESSED_DIR]:
    d.mkdir(parents=True, exist_ok=True)

print(f"📁 RAG folder ready: {RAG_DIR}")

# COMMAND ----------

# ============================================================
# CELL 4 | DATABASE MANAGER
# Handles all read/write operations
# ============================================================

class DatabaseManager:
    """
    Single responsibility: talk to SQLite database.
    All other classes use this — nothing talks to DB directly.
    """

    def __init__(self, db_path=str(DB_PATH)):
        self.db_path = db_path
        self.logger  = self._make_logger()
        self._build_schema()
        print(f"✅ DatabaseManager ready → {self.db_path}")

    # ── Logger ─────────────────────────────────────────────

    def _make_logger(self):
        log_file = LOG_DIR / f"monitor_{datetime.now():%Y%m%d}.log"
        logger   = logging.getLogger("IndiaMonitor")
        logger.setLevel(logging.INFO)
        if not logger.handlers:
            fh = logging.FileHandler(log_file)
            fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
            logger.addHandler(fh)
        return logger

    # ── Connection ─────────────────────────────────────────

    @contextmanager
    def connection(self):
        """Safe auto-commit / auto-rollback connection"""
        conn = sqlite3.connect(
            self.db_path,
            detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
        )
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        try:
            yield conn
            conn.commit()
        except Exception as e:
            conn.rollback()
            self.logger.error(f"DB error: {e}")
            raise
        finally:
            conn.close()

    # ── Schema ─────────────────────────────────────────────

    def _build_schema(self):
        schema = """
        CREATE TABLE IF NOT EXISTS scrape_sessions (
            session_id       INTEGER PRIMARY KEY AUTOINCREMENT,
            session_type     TEXT    NOT NULL,
            started_at       TEXT    NOT NULL,
            completed_at     TEXT,
            status           TEXT    DEFAULT 'running',
            cities_fetched   INTEGER DEFAULT 0,
            articles_found   INTEGER DEFAULT 0,
            error_message    TEXT
        );

        CREATE TABLE IF NOT EXISTS weather_readings (
            reading_id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id          INTEGER,
            city_name           TEXT    NOT NULL,
            state               TEXT,
            recorded_at         TEXT    NOT NULL,
            date_only           TEXT    NOT NULL,
            hour_of_day         INTEGER NOT NULL,
            temperature         REAL,
            feels_like          REAL,
            temp_min            REAL,
            temp_max            REAL,
            humidity            INTEGER,
            pressure            REAL,
            visibility_km       REAL,
            cloudiness          INTEGER,
            weather_main        TEXT,
            weather_description TEXT,
            wind_speed          REAL,
            wind_direction      INTEGER,
            rain_1h             REAL DEFAULT 0,
            snow_1h             REAL DEFAULT 0,
            data_source         TEXT DEFAULT 'OpenWeatherMap',
            UNIQUE(city_name, date_only, hour_of_day)
        );

        CREATE TABLE IF NOT EXISTS strike_news (
            news_id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id       INTEGER,
            title            TEXT    NOT NULL,
            description      TEXT,
            url              TEXT,
            source_name      TEXT,
            keyword_matched  TEXT,
            city_mentioned   TEXT,
            state_mentioned  TEXT,
            severity         TEXT,
            strike_type      TEXT,
            published_date   TEXT,
            scraped_at       TEXT    NOT NULL,
            UNIQUE(title, source_name)
        );

        CREATE TABLE IF NOT EXISTS weather_alerts (
            alert_id      INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id    INTEGER,
            city_name     TEXT NOT NULL,
            state         TEXT,
            alert_type    TEXT NOT NULL,
            alert_message TEXT NOT NULL,
            severity      TEXT,
            triggered_at  TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS daily_summaries (
            summary_id          INTEGER PRIMARY KEY AUTOINCREMENT,
            summary_date        TEXT NOT NULL,
            city_name           TEXT NOT NULL,
            state               TEXT,
            avg_temperature     REAL,
            max_temperature     REAL,
            min_temperature     REAL,
            avg_humidity        REAL,
            total_rainfall      REAL DEFAULT 0,
            avg_wind_speed      REAL,
            dominant_weather    TEXT,
            strike_count        INTEGER DEFAULT 0,
            high_severity_count INTEGER DEFAULT 0,
            weather_alert_count INTEGER DEFAULT 0,
            updated_at          TEXT,
            UNIQUE(summary_date, city_name)
        );

        CREATE INDEX IF NOT EXISTS idx_wr_city_date ON weather_readings(city_name, date_only);
        CREATE INDEX IF NOT EXISTS idx_wr_date      ON weather_readings(date_only);
        CREATE INDEX IF NOT EXISTS idx_sn_date      ON strike_news(published_date);
        CREATE INDEX IF NOT EXISTS idx_sn_city      ON strike_news(city_mentioned);
        CREATE INDEX IF NOT EXISTS idx_ds_date      ON daily_summaries(summary_date);
        """
        with self.connection() as conn:
            conn.executescript(schema)
        self.logger.info("Schema ready")

    # ── Session helpers ────────────────────────────────────

    def session_start(self, session_type="full"):
        with self.connection() as conn:
            cur = conn.execute(
                "INSERT INTO scrape_sessions (session_type, started_at, status) VALUES (?,?,?)",
                (session_type, datetime.now().isoformat(), "running")
            )
            sid = cur.lastrowid
        self.logger.info(f"Session {sid} started [{session_type}]")
        return sid

    def session_end(self, sid, status="success",
                    cities=0, articles=0, error=None):
        with self.connection() as conn:
            conn.execute("""
                UPDATE scrape_sessions
                SET completed_at=?, status=?, cities_fetched=?,
                    articles_found=?, error_message=?
                WHERE session_id=?
            """, (datetime.now().isoformat(), status, cities, articles, error, sid))
        self.logger.info(f"Session {sid} ended [{status}]")

    # ── Write: weather ─────────────────────────────────────

    def write_weather(self, records: list, session_id: int) -> tuple:
        """
        Insert weather records. Skips duplicates silently.
        Returns (saved, skipped)
        """
        saved = skipped = 0
        with self.connection() as conn:
            for r in records:
                try:
                    # Handle both timestamp formats
                    ts_raw = r.get("recorded_at") or r.get("timestamp") or datetime.now().isoformat()
                    try:    ts = datetime.strptime(str(ts_raw), "%Y-%m-%d %H:%M:%S")
                    except: ts = datetime.fromisoformat(str(ts_raw)[:19])

                    conn.execute("""
                        INSERT OR IGNORE INTO weather_readings (
                            session_id, city_name, state,
                            recorded_at, date_only, hour_of_day,
                            temperature, feels_like, temp_min, temp_max,
                            humidity, pressure, visibility_km, cloudiness,
                            weather_main, weather_description,
                            wind_speed, wind_direction,
                            rain_1h, snow_1h, data_source
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        session_id,
                        r.get("city_name") or r.get("city"),
                        r.get("state"),
                        ts.isoformat(),
                        ts.strftime("%Y-%m-%d"),
                        ts.hour,
                        r.get("temperature"),
                        r.get("feels_like"),
                        r.get("temp_min"),
                        r.get("temp_max"),
                        r.get("humidity"),
                        r.get("pressure"),
                        r.get("visibility_km") or r.get("visibility"),
                        r.get("cloudiness"),
                        r.get("weather_main"),
                        r.get("weather_description") or r.get("description"),
                        r.get("wind_speed"),
                        r.get("wind_direction") or r.get("wind_deg"),
                        r.get("rain_1h", 0),
                        r.get("snow_1h", 0),
                        r.get("data_source", "OpenWeatherMap"),
                    ))
                    chg = conn.execute("SELECT changes()").fetchone()[0]
                    if chg: saved += 1
                    else:   skipped += 1
                except Exception as e:
                    self.logger.error(f"write_weather error [{r.get('city_name')}]: {e}")
        return saved, skipped

    # ── Write: alerts ──────────────────────────────────────

    def write_alerts(self, alerts: list, session_id: int) -> int:
        saved = 0
        _type_map = {
            "RAIN": "HEAVY_RAIN", "HEAT": "EXTREME_HEAT",
            "WIND": "STRONG_WIND", "VISIB": "LOW_VISIBILITY",
            "STORM": "THUNDERSTORM",
        }
        with self.connection() as conn:
            for city_alert in alerts:
                for msg in city_alert.get("alerts", []):
                    atype = next(
                        (v for k, v in _type_map.items() if k in msg),
                        "UNKNOWN"
                    )
                    conn.execute("""
                        INSERT INTO weather_alerts
                        (session_id, city_name, state, alert_type,
                         alert_message, severity, triggered_at)
                        VALUES (?,?,?,?,?,?,?)
                    """, (
                        session_id,
                        city_alert["city"],
                        city_alert.get("state", ""),
                        atype, msg,
                        "HIGH" if "EXTREME" in msg else "MEDIUM",
                        datetime.now().isoformat()
                    ))
                    saved += 1
        return saved

    # ── Write: strikes ─────────────────────────────────────

    def write_strikes(self, articles: list, session_id: int) -> tuple:
        saved = skipped = 0
        with self.connection() as conn:
            for a in articles:
                try:
                    sev = str(a.get("severity", "LOW"))
                    for emoji in ["🔴 ", "🟡 ", "🟢 "]:
                        sev = sev.replace(emoji, "")

                    pub_date = None
                    raw = a.get("published_date") or a.get("published", "")
                    if raw:
                        try:    pub_date = pd.to_datetime(raw).strftime("%Y-%m-%d")
                        except: pub_date = datetime.now().strftime("%Y-%m-%d")

                    conn.execute("""
                        INSERT OR IGNORE INTO strike_news (
                            session_id, title, description, url,
                            source_name, keyword_matched,
                            city_mentioned, state_mentioned,
                            severity, strike_type,
                            published_date, scraped_at
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        session_id,
                        str(a.get("title", ""))[:500],
                        str(a.get("description", ""))[:2000],
                        str(a.get("url", ""))[:500],
                        str(a.get("source_name") or a.get("source", "Unknown"))[:100],
                        str(a.get("keyword_matched") or a.get("keyword", ""))[:100],
                        str(a.get("city_mentioned") or a.get("city", "Unknown"))[:100],
                        str(a.get("state_mentioned", ""))[:100],
                        sev,
                        str(a.get("strike_type", "general"))[:50],
                        pub_date,
                        datetime.now().isoformat()
                    ))
                    chg = conn.execute("SELECT changes()").fetchone()[0]
                    if chg: saved += 1
                    else:   skipped += 1
                except Exception as e:
                    self.logger.error(f"write_strikes error: {e}")
        return saved, skipped

    # ── Write: daily summary ───────────────────────────────

    def write_daily_summary(self, date: str = None) -> int:
        date  = date or datetime.now().strftime("%Y-%m-%d")
        count = 0
        with self.connection() as conn:
            cities = conn.execute(
                "SELECT DISTINCT city_name, state FROM weather_readings WHERE date_only=?",
                (date,)
            ).fetchall()

            for row in cities:
                city, state = row["city_name"], row["state"]

                ws = conn.execute("""
                    SELECT ROUND(AVG(temperature),2) avg_t,
                           ROUND(MAX(temperature),2) max_t,
                           ROUND(MIN(temperature),2) min_t,
                           ROUND(AVG(humidity),2)    avg_h,
                           ROUND(SUM(rain_1h),2)     rain,
                           ROUND(AVG(wind_speed),2)  wind
                    FROM weather_readings WHERE city_name=? AND date_only=?
                """, (city, date)).fetchone()

                dom = conn.execute("""
                    SELECT weather_main FROM weather_readings
                    WHERE city_name=? AND date_only=?
                    GROUP BY weather_main ORDER BY COUNT(*) DESC LIMIT 1
                """, (city, date)).fetchone()

                ss = conn.execute("""
                    SELECT COUNT(*) total,
                           SUM(CASE WHEN severity='HIGH' THEN 1 ELSE 0 END) high
                    FROM strike_news WHERE city_mentioned=? AND published_date=?
                """, (city, date)).fetchone()

                ac = conn.execute("""
                    SELECT COUNT(*) c FROM weather_alerts
                    WHERE city_name=? AND DATE(triggered_at)=?
                """, (city, date)).fetchone()["c"]

                conn.execute("""
                    INSERT INTO daily_summaries (
                        summary_date, city_name, state,
                        avg_temperature, max_temperature, min_temperature,
                        avg_humidity, total_rainfall, avg_wind_speed,
                        dominant_weather, strike_count, high_severity_count,
                        weather_alert_count, updated_at
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(summary_date, city_name) DO UPDATE SET
                        avg_temperature=excluded.avg_temperature,
                        max_temperature=excluded.max_temperature,
                        min_temperature=excluded.min_temperature,
                        avg_humidity=excluded.avg_humidity,
                        total_rainfall=excluded.total_rainfall,
                        avg_wind_speed=excluded.avg_wind_speed,
                        dominant_weather=excluded.dominant_weather,
                        strike_count=excluded.strike_count,
                        high_severity_count=excluded.high_severity_count,
                        weather_alert_count=excluded.weather_alert_count,
                        updated_at=excluded.updated_at
                """, (
                    date, city, state,
                    ws["avg_t"], ws["max_t"], ws["min_t"],
                    ws["avg_h"], ws["rain"] or 0, ws["wind"],
                    dom["weather_main"] if dom else None,
                    ss["total"] if ss else 0,
                    ss["high"]  if ss else 0,
                    ac, datetime.now().isoformat()
                ))
                count += 1
        return count

    # ── Read: weather ──────────────────────────────────────

    def read_weather(self, date: str, city: str = None) -> pd.DataFrame:
        q = "SELECT * FROM weather_readings WHERE date_only=?"
        p = [date]
        if city:
            q += " AND city_name=?"
            p.append(city)
        q += " ORDER BY city_name, hour_of_day"
        with self.connection() as conn:
            rows = conn.execute(q, p).fetchall()
        return pd.DataFrame([dict(r) for r in rows])

    def read_weather_range(self, city: str, start: str, end: str) -> pd.DataFrame:
        with self.connection() as conn:
            rows = conn.execute("""
                SELECT * FROM weather_readings
                WHERE city_name=? AND date_only BETWEEN ? AND ?
                ORDER BY recorded_at
            """, (city, start, end)).fetchall()
        return pd.DataFrame([dict(r) for r in rows])

    def read_latest_all_cities(self) -> pd.DataFrame:
        today = datetime.now().strftime("%Y-%m-%d")
        with self.connection() as conn:
            rows = conn.execute("""
                SELECT w.* FROM weather_readings w
                INNER JOIN (
                    SELECT city_name, MAX(recorded_at) latest
                    FROM weather_readings WHERE date_only=?
                    GROUP BY city_name
                ) lw ON w.city_name=lw.city_name AND w.recorded_at=lw.latest
                ORDER BY w.temperature DESC
            """, (today,)).fetchall()
        return pd.DataFrame([dict(r) for r in rows])

    # ── Read: strikes ──────────────────────────────────────

    def read_strikes(self, date: str = None, city: str = None,
                     severity: str = None, strike_type: str = None,
                     start: str = None, end: str = None) -> pd.DataFrame:
        q, p = "SELECT * FROM strike_news WHERE 1=1", []
        if date:         q += " AND published_date=?";           p.append(date)
        if start:        q += " AND published_date>=?";          p.append(start)
        if end:          q += " AND published_date<=?";          p.append(end)
        if city:         q += " AND city_mentioned=?";           p.append(city)
        if severity:     q += " AND severity=?";                 p.append(severity.upper())
        if strike_type:  q += " AND strike_type=?";              p.append(strike_type)
        q += " ORDER BY published_date DESC"
        with self.connection() as conn:
            rows = conn.execute(q, p).fetchall()
        return pd.DataFrame([dict(r) for r in rows])

    def read_strikes_search(self, keyword: str) -> pd.DataFrame:
        with self.connection() as conn:
            rows = conn.execute("""
                SELECT * FROM strike_news
                WHERE title LIKE ? OR description LIKE ?
                ORDER BY published_date DESC
            """, (f"%{keyword}%", f"%{keyword}%")).fetchall()
        return pd.DataFrame([dict(r) for r in rows])

    # ── Read: summaries ────────────────────────────────────

    def read_summary(self, start: str, end: str = None,
                     city: str = None) -> pd.DataFrame:
        end = end or datetime.now().strftime("%Y-%m-%d")
        q, p = "SELECT * FROM daily_summaries WHERE summary_date BETWEEN ? AND ?", [start, end]
        if city:
            q += " AND city_name=?"
            p.append(city)
        q += " ORDER BY summary_date DESC, city_name"
        with self.connection() as conn:
            rows = conn.execute(q, p).fetchall()
        return pd.DataFrame([dict(r) for r in rows])

    # ── Read: stats ────────────────────────────────────────

    def read_stats(self) -> dict:
        with self.connection() as conn:
            return {
                "weather_records" : conn.execute("SELECT COUNT(*) FROM weather_readings").fetchone()[0],
                "strike_articles" : conn.execute("SELECT COUNT(*) FROM strike_news").fetchone()[0],
                "weather_alerts"  : conn.execute("SELECT COUNT(*) FROM weather_alerts").fetchone()[0],
                "sessions_run"    : conn.execute("SELECT COUNT(*) FROM scrape_sessions").fetchone()[0],
                "cities_tracked"  : conn.execute("SELECT COUNT(DISTINCT city_name) FROM weather_readings").fetchone()[0],
                "date_from"       : conn.execute("SELECT MIN(date_only) FROM weather_readings").fetchone()[0],
                "date_to"         : conn.execute("SELECT MAX(date_only) FROM weather_readings").fetchone()[0],
                "db_size_mb"      : round(os.path.getsize(self.db_path) / 1024 / 1024, 3),
            }

    # ── Export ─────────────────────────────────────────────

    def export(self, date: str = None, fmt: str = "csv") -> dict:
        date  = date or datetime.now().strftime("%Y-%m-%d")
        tag   = date.replace("-", "")
        files = {}
        data  = {
            "weather" : self.read_weather(date),
            "strikes" : self.read_strikes(date=date),
            "summary" : self.read_summary(date, date),
        }
        for name, df in data.items():
            if df.empty:
                continue
            if fmt == "csv":
                p = CSV_DIR / f"{name}_{tag}.csv"
                df.to_csv(p, index=False)
            else:
                p = JSON_DIR / f"{name}_{tag}.json"
                payload = {
                    "date": date,
                    "generated_at": datetime.now().isoformat(),
                    "count": len(df),
                    "data": df.to_dict(orient="records"),
                }
                with open(p, "w") as f:
                    json.dump(payload, f, indent=2, default=str)
            files[name] = str(p)
        return files
        print(f"  📄 {p}")


# ── Initialise ─────────────────────────────────────────────
db = DatabaseManager()
print("✅ Cell 4 complete")

# COMMAND ----------

# ============================================================
# CELL 5 | WEATHER SERVICE
# Two sources:
#   - OpenWeatherMap  → current / live
#   - Open-Meteo      → historical (free, no key)
# ============================================================

class WeatherService:
    """Fetches weather data from appropriate API based on date"""

    OWM_URL        = "https://api.openweathermap.org/data/2.5/weather"
    METEO_HIST_URL = "https://archive-api.open-meteo.com/v1/archive"
    METEO_NOW_URL  = "https://api.open-meteo.com/v1/forecast"

    WEATHER_CODES = {
        0:"clear sky",    1:"mainly clear",  2:"partly cloudy",
        3:"overcast",    45:"fog",           48:"icy fog",
        51:"light drizzle", 53:"moderate drizzle", 55:"heavy drizzle",
        61:"light rain", 63:"moderate rain", 65:"heavy rain",
        71:"light snow", 73:"moderate snow", 75:"heavy snow",
        80:"light showers", 81:"moderate showers", 82:"heavy showers",
        95:"thunderstorm", 96:"thunderstorm with hail",
    }
    WEATHER_MAIN = {
        **{k:"Clear"        for k in [0,1]},
        **{k:"Clouds"       for k in [2,3]},
        **{k:"Fog"          for k in [45,48]},
        **{k:"Drizzle"      for k in [51,53,55]},
        **{k:"Rain"         for k in [61,63,65,80,81,82]},
        **{k:"Snow"         for k in [71,73,75]},
        **{k:"Thunderstorm" for k in [95,96]},
    }

    def __init__(self, api_key: str, cities: dict):
        self.api_key = api_key
        self.cities  = cities

    # ── Public API ─────────────────────────────────────────

    def fetch_current(self) -> list:
        """Fetch live weather for all cities via OpenWeatherMap"""
        print("🌤️  Fetching current weather (OpenWeatherMap)...")
        results = []
        for city, coords in self.cities.items():
            print(f"   📍 {city:<15}", end=" ")
            rec = self._owm_one(city, coords)
            if rec:
                results.append(rec)
                print(f"✅ {rec['temperature']:.1f}°C  {rec['weather_description']}")
            else:
                print("❌ failed")
            time.sleep(0.5)
        print(f"   → {len(results)}/{len(self.cities)} cities fetched")
        return results

    def fetch_historical(self, date: str, city: str = None) -> list:
        """Fetch historical weather via Open-Meteo (free, no key)"""
        cities = {city: self.cities[city]} if city else self.cities
        today  = datetime.now().strftime("%Y-%m-%d")
        url    = self.METEO_HIST_URL if date < today else self.METEO_NOW_URL

        print(f"   📡 Open-Meteo {'archive' if date < today else 'forecast'} API")
        results = []
        for city_name, coords in cities.items():
            print(f"   📍 {city_name}...", end=" ")
            recs = self._meteo_one(city_name, coords, date, url)
            results.extend(recs)
            print(f"✅ {len(recs)} hourly readings")
            time.sleep(0.3)
        return results

    def get_alerts(self, records: list) -> list:
        """Flag dangerous weather conditions"""
        alerts, t = [], ALERT_THRESHOLDS
        city_latest = {}
        for r in records:
            city_latest[r.get("city_name") or r.get("city")] = r

        for city, r in city_latest.items():
            msgs = []
            if (r.get("rain_1h") or 0)    > t["rain_mm_per_hr"]: msgs.append(f"🌧️  HEAVY RAIN {r['rain_1h']}mm/hr")
            if (r.get("temperature") or 0) > t["temp_extreme_c"]: msgs.append(f"🌡️  EXTREME HEAT {r['temperature']}°C")
            if (r.get("wind_speed") or 0)  > t["wind_ms"]:        msgs.append(f"💨 STRONG WIND {r['wind_speed']}m/s")
            if 0 < (r.get("visibility_km") or 999) < t["visibility_km"]: msgs.append(f"🌫️  LOW VISIBILITY {r['visibility_km']}km")
            if "thunderstorm" in str(r.get("weather_main","")).lower():   msgs.append("⛈️  THUNDERSTORM")
            if msgs:
                alerts.append({"city": city, "state": r.get("state",""), "alerts": msgs})
        return alerts

    # ── Private helpers ────────────────────────────────────

    def _owm_one(self, city: str, coords: dict) -> dict | None:
        try:
            resp = requests.get(self.OWM_URL, params={
                "lat": coords["lat"], "lon": coords["lon"],
                "appid": self.api_key, "units": "metric"
            }, timeout=10)
            resp.raise_for_status()
            d = resp.json()
            now = datetime.now()
            return {
                "city_name"          : city,
                "state"              : coords["state"],
                "recorded_at"        : now.strftime("%Y-%m-%d %H:%M:%S"),
                "date_only"          : now.strftime("%Y-%m-%d"),
                "hour_of_day"        : now.hour,
                "temperature"        : d["main"]["temp"],
                "feels_like"         : d["main"]["feels_like"],
                "temp_min"           : d["main"]["temp_min"],
                "temp_max"           : d["main"]["temp_max"],
                "humidity"           : d["main"]["humidity"],
                "pressure"           : d["main"]["pressure"],
                "weather_main"       : d["weather"][0]["main"],
                "weather_description": d["weather"][0]["description"],
                "wind_speed"         : d["wind"]["speed"],
                "wind_direction"     : d["wind"].get("deg", 0),
                "visibility_km"      : round(d.get("visibility", 0) / 1000, 2),
                "cloudiness"         : d["clouds"]["all"],
                "rain_1h"            : d.get("rain", {}).get("1h", 0),
                "snow_1h"            : 0,
                "data_source"        : "OpenWeatherMap",
            }
        except Exception as e:
            print(f"❌ OWM error [{city}]: {e}")
            return None

    def _meteo_one(self, city: str, coords: dict, date: str, url: str) -> list:
        try:
            resp = requests.get(url, params={
                "latitude": coords["lat"], "longitude": coords["lon"],
                "start_date": date, "end_date": date,
                "hourly": [
                    "temperature_2m", "relative_humidity_2m",
                    "apparent_temperature", "rain", "snowfall",
                    "weather_code", "surface_pressure",
                    "cloud_cover", "visibility",
                    "wind_speed_10m", "wind_direction_10m",
                ],
                "timezone": "Asia/Kolkata",
                "wind_speed_unit": "ms",
            }, timeout=15)
            resp.raise_for_status()
            h = resp.json().get("hourly", {})
            times = h.get("time", [])

            records = []
            temps   = h.get("temperature_2m", [None]*len(times))
            for i, ts in enumerate(times):
                code  = (h.get("weather_code") or [None]*len(times))[i]
                records.append({
                    "city_name"          : city,
                    "state"              : coords["state"],
                    "recorded_at"        : ts,
                    "date_only"          : date,
                    "hour_of_day"        : int(ts[11:13]),
                    "temperature"        : temps[i],
                    "feels_like"         : (h.get("apparent_temperature")  or [None]*len(times))[i],
                    "temp_min"           : min(t for t in temps if t is not None),
                    "temp_max"           : max(t for t in temps if t is not None),
                    "humidity"           : (h.get("relative_humidity_2m")  or [None]*len(times))[i],
                    "pressure"           : (h.get("surface_pressure")      or [None]*len(times))[i],
                    "visibility_km"      : round(((h.get("visibility") or [0]*len(times))[i] or 0)/1000, 2),
                    "cloudiness"         : (h.get("cloud_cover")           or [None]*len(times))[i],
                    "weather_main"       : self.WEATHER_MAIN.get(code, "Unknown"),
                    "weather_description": self.WEATHER_CODES.get(code, f"code {code}"),
                    "wind_speed"         : (h.get("wind_speed_10m")        or [None]*len(times))[i],
                    "wind_direction"     : (h.get("wind_direction_10m")    or [None]*len(times))[i],
                    "rain_1h"            : (h.get("rain")                  or [0]*len(times))[i] or 0,
                    "snow_1h"            : (h.get("snowfall")              or [0]*len(times))[i] or 0,
                    "data_source"        : "Open-Meteo",
                })
            return records
        except Exception as e:
            print(f"❌ Meteo error [{city}]: {e}")
            return []


weather_svc = WeatherService(OPENWEATHER_API_KEY, INDIA_CITIES)
print("✅ Cell 5 complete — WeatherService ready")

# COMMAND ----------

# ============================================================
# CELL 6 | NEWS SERVICE
# Scrapes Google News RSS for strike articles
# No API key needed
# ============================================================

class NewsService:
    """Scrapes and classifies transportation strike news"""

    RSS_URL = "https://news.google.com/rss/search"

    _SEVERITY_HIGH   = ["bharat bandh","national strike","indefinite","complete shutdown"]
    _SEVERITY_MEDIUM = ["state bandh","24-hour","48-hour","city strike"]

    _TYPE_MAP = [
        (["bus","rtc","ksrtc"],          "bus"),
        (["truck","lorry"],              "truck"),
        (["rail","train"],               "railway"),
        (["auto","rickshaw"],            "auto"),
        (["taxi","cab"],                 "taxi"),
        (["metro"],                      "metro"),
        (["bandh","hartal"],             "bandh"),
    ]

    def __init__(self, keywords: list, cities: dict):
        self.keywords = keywords
        self.cities   = cities

    # ── Public API ─────────────────────────────────────────

    def fetch(self, date: str = None, city: str = None) -> list:
        """
        Fetch strike news.
        date → filter to specific date (uses after:/before: operators)
        city → filter to specific city
        """
        queries = self._build_queries(city)
        raw     = []
        seen    = set()

        date_filter = self._date_filter(date) if date else ""

        for q in queries:
            full_query = f"{q} {date_filter}".strip()
            articles   = self._rss_search(full_query)
            for a in articles:
                key = a["title"].lower()[:80]
                if key not in seen:
                    seen.add(key)
                    a["keyword_matched"] = q
                    a["city_mentioned"]  = self._detect_city(a["title"])
                    a["state_mentioned"] = INDIA_CITIES.get(a["city_mentioned"], {}).get("state", "")
                    a["severity"]        = self._severity(a["title"])
                    a["strike_type"]     = self._classify(a["title"])
                    raw.append(a)
            time.sleep(0.4)

        # Filter by city if requested
        if city:
            raw = [
                a for a in raw
                if a["city_mentioned"] == city
                or city.lower() in a["title"].lower()
            ]

        print(f"   → {len(raw)} unique articles found")
        return raw

    # ── Private helpers ────────────────────────────────────

    def _build_queries(self, city: str = None) -> list:
        queries = [f"{kw} India" for kw in self.keywords]
        if city:
            queries += [f"transport strike {city}", f"bandh {city}"]
        else:
            queries += [f"transport strike {c}" for c in list(self.cities)[:10]]
        return queries

    def _date_filter(self, date: str) -> str:
        try:
            dt = datetime.strptime(date, "%Y-%m-%d")
            return f"after:{(dt - timedelta(days=1)):%Y-%m-%d} before:{(dt + timedelta(days=1)):%Y-%m-%d}"
        except:
            return ""

    def _rss_search(self, query: str) -> list:
        articles = []
        try:
            url  = f"{self.RSS_URL}?q={requests.utils.quote(query)}&hl=en-IN&gl=IN&ceid=IN:en"
            resp = requests.get(url, headers={"User-Agent": random_ua()}, timeout=10)
            soup = BeautifulSoup(resp.content, "xml")
            for item in soup.find_all("item")[:8]:
                title = item.find("title")
                if not title: continue
                desc  = item.find("description")
                link  = item.find("link")
                pub   = item.find("pubDate")
                src   = item.find("source")
                articles.append({
                    "title"        : title.get_text(strip=True)[:500],
                    "description"  : (desc.get_text(strip=True) if desc else "")[:1000],
                    "url"          : (link.get_text(strip=True) if link else ""),
                    "source_name"  : (src.get_text(strip=True) if src else "Google News"),
                    "published_date": self._parse_date(pub.get_text(strip=True) if pub else ""),
                    "scraped_at"   : datetime.now().isoformat(),
                })
        except Exception as e:
            pass  # Silent fail — RSS is best-effort
        return articles

    def _parse_date(self, raw: str) -> str:
        for fmt in ["%a, %d %b %Y %H:%M:%S %Z", "%a, %d %b %Y %H:%M:%S %z",
                    "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d"]:
            try: return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
            except: continue
        return datetime.now().strftime("%Y-%m-%d")

    def _detect_city(self, text: str) -> str:
        t = text.lower()
        for city in self.cities:
            if city.lower() in t:
                return city
        return "National/Unknown"

    def _severity(self, title: str) -> str:
        t = title.lower()
        if any(w in t for w in self._SEVERITY_HIGH):   return "HIGH"
        if any(w in t for w in self._SEVERITY_MEDIUM): return "MEDIUM"
        return "LOW"

    def _classify(self, text: str) -> str:
        t = text.lower()
        for keywords, label in self._TYPE_MAP:
            if any(w in t for w in keywords):
                return label
        return "general"


news_svc = NewsService(STRIKE_KEYWORDS, INDIA_CITIES)
print("✅ Cell 6 complete — NewsService ready")

# COMMAND ----------

# ============================================================
# CELL 7 | SMART FETCHER
# Single entry point for all data requests
# DB first → API fallback → save → return
# ============================================================

class SmartFetcher:
    """
    The only class you interact with directly.

    Usage:
        result = fetcher.get("2024-01-15")
        result = fetcher.get("2024-01-15", city="Mumbai")
        result = fetcher.get_range("2024-01-10", "2024-01-15")
    """

    def __init__(self, db: DatabaseManager,
                 weather: WeatherService,
                 news: NewsService):
        self.db      = db
        self.weather = weather
        self.news    = news
        print("✅ SmartFetcher ready")

    # ── Main entry point ───────────────────────────────────

    def get(self, date: str, city: str = None) -> dict:
        """
        Get weather + strikes for any date.

        Checks DB first.
        If missing → fetches from API → saves to DB → returns.

        Parameters:
            date  → "YYYY-MM-DD"
            city  → city name or None for all

        Returns dict:
            {
                date     : str,
                city     : str,
                source   : "database" | "api" | "mixed",
                weather  : DataFrame,
                strikes  : DataFrame,
            }
        """
        self._validate_date(date)

        print(f"\n{'='*52}")
        print(f"  📅 Date   : {date}")
        print(f"  🏙️  City   : {city or 'All Cities'}")
        print(f"{'='*52}")

        # ── Step 1: DB check ───────────────────────────────
        print("\n🔍 Checking database...")
        w_df = self.db.read_weather(date, city)
        s_df = self.db.read_strikes(date=date, city=city)

        w_found = not w_df.empty
        s_found = not s_df.empty
        print(f"   Weather : {'✅ found ({} records)'.format(len(w_df)) if w_found else '❌ not found'}")
        print(f"   Strikes : {'✅ found ({} articles)'.format(len(s_df)) if s_found else '❌ not found'}")

        # ── Step 2: API fetch if missing ───────────────────
        if not w_found:
            print("\n🌐 Fetching weather from API...")
            records = self.weather.fetch_historical(date, city)
            if records:
                sid = self.db.session_start("api_weather")
                saved, skip = self.db.write_weather(records, sid)
                self.db.session_end(sid, cities=len(set(r["city_name"] for r in records)))
                self.db.write_daily_summary(date)
                print(f"   💾 {saved} saved | {skip} skipped")
                w_df = self.db.read_weather(date, city)

        if not s_found:
            print("\n📰 Fetching strikes from News API...")
            articles = self.news.fetch(date=date, city=city)
            if articles:
                sid = self.db.session_start("api_news")
                saved, skip = self.db.write_strikes(articles, sid)
                self.db.session_end(sid, articles=len(articles))
                print(f"   💾 {saved} saved | {skip} skipped")
                s_df = self.db.read_strikes(date=date, city=city)

        # ── Step 3: Return ─────────────────────────────────
        source = (
            "database" if (w_found and s_found) else
            "mixed"    if (w_found or s_found)  else
            "api"
        )

        print(f"\n{'='*52}")
        print(f"  ✅ Source  : {source}")
        print(f"  🌤️  Weather : {len(w_df)} records")
        print(f"  📰 Strikes : {len(s_df)} articles")
        print(f"{'='*52}")

        return {
            "date"    : date,
            "city"    : city or "All Cities",
            "source"  : source,
            "weather" : w_df,
            "strikes" : s_df,
        }

    def get_range(self, start: str, end: str, city: str = None) -> dict:
        """
        Get data for a date range.
        Each date checked individually (DB → API).

        Returns dict:
            {
                weather  : combined DataFrame,
                strikes  : combined DataFrame,
                summary  : per-date summary DataFrame,
            }
        """
        start_dt = datetime.strptime(start, "%Y-%m-%d")
        end_dt   = datetime.strptime(end,   "%Y-%m-%d")
        days     = (end_dt - start_dt).days + 1

        print(f"\n📅 Fetching {days} days: {start} → {end}")
        print(f"🏙️  City: {city or 'All Cities'}")

        all_w, all_s, summary = [], [], []
        current = start_dt

        while current <= end_dt:
            date_str = current.strftime("%Y-%m-%d")
            result   = self.get(date_str, city)

            if not result["weather"].empty: all_w.append(result["weather"])
            if not result["strikes"].empty: all_s.append(result["strikes"])

            summary.append({
                "date"    : date_str,
                "source"  : result["source"],
                "weather" : len(result["weather"]),
                "strikes" : len(result["strikes"]),
            })
            current += timedelta(days=1)
            time.sleep(0.3)

        w_combined = pd.concat(all_w, ignore_index=True) if all_w else pd.DataFrame()
        s_combined = pd.concat(all_s, ignore_index=True) if all_s else pd.DataFrame()
        summary_df = pd.DataFrame(summary)

        print(f"\n✅ Range complete")
        print(f"   Weather records : {len(w_combined)}")
        print(f"   Strike articles : {len(s_combined)}")
        print(summary_df.to_string(index=False))

        return {"weather": w_combined, "strikes": s_combined, "summary": summary_df}

    def _validate_date(self, date: str):
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            raise ValueError(f"Invalid date '{date}' — use YYYY-MM-DD")
        if date > datetime.now().strftime("%Y-%m-%d"):
            raise ValueError(f"Cannot fetch future date: {date}")


fetcher = SmartFetcher(db, weather_svc, news_svc)
print("✅ Cell 7 complete")

# COMMAND ----------

# ============================================================
# CELL 8 | PIPELINE
# Collects TODAY's live data and saves to DB
# Run this daily / on a schedule
# ============================================================

def run_pipeline():
    """Collect live weather + strike news for today"""

    print("=" * 52)
    print("🚀 INDIA MONITOR PIPELINE")
    print(f"   {datetime.now():%Y-%m-%d %H:%M:%S}")
    print("=" * 52)

    sid = db.session_start("pipeline")

    try:
        # ── Weather ────────────────────────────────────────
        print("\n[1/4] Fetching live weather...")
        records = weather_svc.fetch_current()
        saved, skip = db.write_weather(records, sid)
        print(f"      💾 {saved} saved | {skip} skipped")

        # ── Alerts ─────────────────────────────────────────
        print("\n[2/4] Checking weather alerts...")
        alerts = weather_svc.get_alerts(records)
        a_saved = db.write_alerts(alerts, sid)
        print(f"      💾 {a_saved} alerts saved")
        if alerts:
            for a in alerts:
                print(f"      ⚠️  {a['city']}: {', '.join(a['alerts'])}")

        # ── News ───────────────────────────────────────────
        print("\n[3/4] Scraping strike news...")
        articles = news_svc.fetch()
        n_saved, n_skip = db.write_strikes(articles, sid)
        print(f"      💾 {n_saved} saved | {n_skip} skipped")

        # ── Summary ────────────────────────────────────────
        print("\n[4/4] Computing daily summary...")
        count = db.write_daily_summary()
        print(f"      📊 {count} cities summarised")

        # ── Export ─────────────────────────────────────────
        print("\n📤 Exporting CSV + JSON...")
        db.export(fmt="csv")
        db.export(fmt="json")

        db.session_end(sid, cities=len(records), articles=len(articles))

        print("\n" + "=" * 52)
        print("✅ PIPELINE COMPLETE")
        print(f"   Weather : {len(records)} cities")
        print(f"   Strikes : {len(articles)} articles")
        print(f"   Alerts  : {len(alerts)} cities flagged")
        print("=" * 52)

    except Exception as e:
        db.session_end(sid, status="failed", error=str(e))
        print(f"\n❌ Pipeline failed: {e}")
        import traceback; traceback.print_exc()


# ── Run it ─────────────────────────────────────────────────
run_pipeline()

# COMMAND ----------

# ============================================================
# CELL 9 | QUERY INTERFACE
# All the ways to access your data
# ============================================================

# ── Helpers ────────────────────────────────────────────────
TODAY     = datetime.now().strftime("%Y-%m-%d")
YESTERDAY = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
LAST_7    = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
LAST_30   = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

# ============================================================
# EXAMPLES — uncomment what you need
# ============================================================

# ── DB overview ────────────────────────────────────────────
stats = db.read_stats()
print("📊 Database Stats:")
for k, v in stats.items():
    print(f"   {k:<20}: {v}")

# ── Today's weather ────────────────────────────────────────
# today_weather = db.read_latest_all_cities()
# print(today_weather[["city_name","temperature","humidity","weather_description"]])

# ── Any date (DB → API auto) ───────────────────────────────
# result = fetcher.get("2024-01-15")
# result = fetcher.get("2024-01-15", city="Mumbai")
# print(result["weather"])
# print(result["strikes"])

# ── Date range ─────────────────────────────────────────────
# result = fetcher.get_range("2024-01-10", "2024-01-15")
# result = fetcher.get_range("2024-01-10", "2024-01-15", city="Delhi")

# ── Weather history ────────────────────────────────────────
# df = db.read_weather_range("Mumbai", LAST_30, TODAY)

# ── Strike queries ─────────────────────────────────────────
# df = db.read_strikes(severity="HIGH")
# df = db.read_strikes(city="Mumbai", start=LAST_30)
# df = db.read_strikes(strike_type="bus")
# df = db.read_strikes_search("bandh")

# ── Daily summary ──────────────────────────────────────────
# df = db.read_summary(LAST_7)
# df = db.read_summary(LAST_30, city="Delhi")

# ── Export specific date ───────────────────────────────────
# db.export("2024-01-15", fmt="csv")
# db.export("2024-01-15", fmt="json")

print("\n✅ Query interface ready")
print("   Uncomment any example above to run it")

# COMMAND ----------

# ============================================================
# CELL 10 | RAG INSTALLATION
# ============================================================

import subprocess, sys

RAG_PACKAGES = [
    "sentence-transformers",   # Local embeddings
    "faiss-cpu",               # Vector store
    "langchain",               # RAG framework
    "langchain-community",     # Document loaders
    "pypdf",                   # PDF support
    "python-docx",             # Word doc support
    "tiktoken",                # Token counting
    "numpy",                   # Vector math
    "scikit-learn",            # ML prep
    "openpyxl",                # Excel support
]

print("📦 Installing RAG packages...")
for pkg in RAG_PACKAGES:
    r = subprocess.run(
        [sys.executable, "-m", "pip", "install", pkg, "-q"],
        capture_output=True, text=True
    )
    print(f"  {'✅' if r.returncode == 0 else '❌'} {pkg}")

print("\n✅ RAG Installation complete!")

# COMMAND ----------

# ============================================================
# CELL 11 | RAG IMPORTS
# ============================================================

import os
import json
import pickle
import hashlib
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple

# Sentence Transformers (local embeddings)
try:
    from sentence_transformers import SentenceTransformer
    print("✅ sentence_transformers loaded")
except ImportError:
    print("❌ sentence_transformers missing - run Cell 10")

# FAISS (vector store)
try:
    import faiss
    print("✅ faiss loaded")
except ImportError:
    print("❌ faiss missing - run Cell 10")

# Document loaders
try:
    import pypdf
    from pypdf import PdfReader
    print("✅ pypdf loaded")
except ImportError:
    print("⚠️  pypdf missing - PDF loading won't work")

try:
    import docx
    print("✅ python-docx loaded")
except ImportError:
    print("⚠️  python-docx missing - Word loading won't work")

try:
    import openpyxl
    print("✅ openpyxl loaded")
except ImportError:
    print("⚠️  openpyxl missing - Excel loading won't work")

print("\n✅ All RAG imports done!")

# COMMAND ----------

# ============================================================
# CELL 12 | RAG CONFIGURATION
# Add this to your existing config from Cell 3
# ============================================================

# ── RAG Folder Structure ────────────────────────────────────
RAG_DIR       = BASE_DIR / "rag"
DOCS_DIR      = RAG_DIR  / "documents"
VECTOR_DIR    = RAG_DIR  / "vector_store"
PROCESSED_DIR = RAG_DIR  / "processed"
CHUNKS_DIR    = RAG_DIR  / "chunks"

for d in [DOCS_DIR, VECTOR_DIR, PROCESSED_DIR, CHUNKS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ── Embedding Model ─────────────────────────────────────────
# Free, runs locally, no API key needed
# Downloads once (~90MB), then works offline
EMBEDDING_MODEL = "all-MiniLM-L6-v2"

# ── Chunking Settings ───────────────────────────────────────
CHUNK_SIZE    = 500    # characters per chunk
CHUNK_OVERLAP = 50     # overlap between chunks

# ── Retrieval Settings ──────────────────────────────────────
TOP_K_RESULTS = 5      # how many chunks to retrieve per query

# ── Document Categories ─────────────────────────────────────
# These are the types of documents RAG understands
DOC_CATEGORIES = {
    "sla"             : "Service Level Agreement",
    "kedb"            : "Known Error Database",
    "resolution_log"  : "Past Resolution Logs",
    "sop"             : "Standard Operating Procedure",
    "policy"          : "Policy Document",
}

print("=" * 52)
print("⚙️  RAG CONFIGURATION")
print("=" * 52)
print(f"  Documents folder : {DOCS_DIR}")
print(f"  Vector store     : {VECTOR_DIR}")
print(f"  Embedding model  : {EMBEDDING_MODEL}")
print(f"  Chunk size       : {CHUNK_SIZE} chars")
print(f"  Chunk overlap    : {CHUNK_OVERLAP} chars")
print(f"  Top K results    : {TOP_K_RESULTS}")
print("=" * 52)
print("\n📂 Drop your documents in:")
print(f"   {DOCS_DIR}")
print("\n   Supported formats:")
print("   .pdf  .docx  .txt  .csv  .xlsx")

# COMMAND ----------

# ============================================================
# CELL 13 | SAMPLE DOCUMENT GENERATOR
# Creates realistic sample documents if you don't have real ones
# Replace with your actual documents later
# ============================================================

def create_sample_documents():
    """
    Creates sample SLA, KEDB, Resolution Log, SOP documents
    in the documents folder.
    Replace these with your real documents.
    """

    # ── 1. SLA Document ────────────────────────────────────
    sla_content = """
SERVICE LEVEL AGREEMENT - INDIA TRANSPORT MONITORING SYSTEM
Version: 2.1 | Effective Date: January 1, 2024

SECTION 1: PURPOSE
This SLA defines the response and resolution commitments
for transport disruption incidents across India.

SECTION 2: INCIDENT PRIORITY LEVELS

P1 - CRITICAL
Definition: Complete transport shutdown affecting 3+ metro cities
            OR Bharat Bandh declared
            OR National highway blockade
Response Time: 30 minutes
Resolution Time: 4 hours
Escalation: Immediate escalation to Regional Director
Notification: SMS + Email to all stakeholders within 15 minutes
Examples: Bharat Bandh, National truck strike, Railway shutdown

P2 - HIGH
Definition: Transport strike affecting 1-2 metro cities
            OR State-wide bandh
            OR Major highway blockade (500+ km)
Response Time: 1 hour
Resolution Time: 8 hours
Escalation: Escalate to City Manager after 2 hours
Notification: Email to stakeholders within 30 minutes
Examples: State bandh, City-wide bus strike, Port blockade

P3 - MEDIUM
Definition: Partial transport disruption in Tier-1 city
            OR Auto/taxi strike in metro
            OR District-level bandh
Response Time: 2 hours
Resolution Time: 24 hours
Escalation: Escalate to Supervisor after 4 hours
Notification: Email within 1 hour
Examples: Auto strike, Cab strike, District bandh

P4 - LOW
Definition: Minor disruption, single route or area
            OR Protest without complete shutdown
Response Time: 4 hours
Resolution Time: 48 hours
Escalation: Standard escalation after 8 hours
Notification: Daily report
Examples: Single route blockade, Partial auto strike

SECTION 3: WEATHER-RELATED SLA ADJUSTMENTS

Extreme Weather Conditions:
- Heavy Rain (>20mm/hr): Response time extended by 50%
- Cyclone Warning: P1 protocol activated automatically
- Flood Alert: Resolution time extended by 100%
- Extreme Heat (>45C): Additional resources deployed

SECTION 4: CITY-SPECIFIC SLA TERMS

Mumbai:
- Local train disruption: Treated as P1 automatically
- Port strike: P2 with maritime authority notification
- Auto strike: P3 with alternate transport activation

Delhi:
- Metro disruption: P1 automatically
- Airport connectivity: P1 automatically
- Ring road blockade: P2

Bangalore:
- Tech corridor disruption: P2 (economic impact)
- Airport road: P2

Chennai:
- Port disruption: P1 (major export hub)
- Auto strike: P3

SECTION 5: ESCALATION MATRIX

Level 1: Field Coordinator (0-2 hours)
Level 2: City Manager (2-4 hours)
Level 3: Regional Director (4-8 hours)
Level 4: National Head (8+ hours)

SECTION 6: PENALTIES

SLA Breach - P1: Rs 50,000 per hour beyond resolution time
SLA Breach - P2: Rs 20,000 per hour beyond resolution time
SLA Breach - P3: Rs 5,000 per hour beyond resolution time

SECTION 7: REPORTING

Daily Report: Sent by 9 AM every day
Weekly Report: Sent every Monday by 10 AM
Monthly Review: First Friday of every month
Incident Report: Within 24 hours of P1/P2 resolution
"""

    # ── 2. KEDB Document ────────────────────────────────────
    kedb_content = """
KNOWN ERROR DATABASE (KEDB)
INDIA TRANSPORT MONITORING SYSTEM
Last Updated: January 2024

═══════════════════════════════════════════════════════
KEDB-001 | Bus Strike - Maharashtra
═══════════════════════════════════════════════════════
Error ID        : KEDB-001
Category        : Bus Strike
Affected Cities : Mumbai, Pune, Nagpur
Trigger         : MSRTC driver wage disputes
Frequency       : 2-3 times per year
Typical Duration: 24-72 hours
Priority Level  : P2

Root Cause:
Maharashtra State Road Transport Corporation (MSRTC)
driver unions periodically strike over wage revision,
pension demands, and service conditions.

Known Workarounds:
1. Activate private bus operator agreements (pre-signed MOUs)
2. Coordinate with Mumbai Metro for increased frequency
3. Deploy app-based cab aggregator emergency fleet
4. Notify corporates for WFH advisories
5. Coordinate with railways for additional local trains

Resolution Steps:
Step 1: Confirm strike via MSRTC official channels
Step 2: Activate backup transport MOU within 1 hour
Step 3: Issue public advisory via social media
Step 4: Monitor situation every 2 hours
Step 5: Coordinate with state transport ministry
Step 6: Update stakeholders every 4 hours

Past Occurrences:
- March 2023: 3-day strike, resolved via wage agreement
- August 2022: 2-day strike, resolved via court order
- January 2022: 1-day strike, resolved via negotiation

═══════════════════════════════════════════════════════
KEDB-002 | Truck Strike - National
═══════════════════════════════════════════════════════
Error ID        : KEDB-002
Category        : Truck/Lorry Strike
Affected Cities : All major cities
Trigger         : Fuel price hike, toll disputes, new regulations
Frequency       : 1-2 times per year
Typical Duration: 3-7 days
Priority Level  : P1

Root Cause:
All India Motor Transport Congress (AIMTC) calls strikes
when fuel prices increase significantly or new regulations
impact profitability. Affects entire supply chain.

Known Workarounds:
1. Activate rail freight for critical goods
2. Coordinate with air cargo for essential items
3. Issue essential services exemption list
4. Deploy government transport for critical supplies
5. Coordinate with state governments for local exemptions

Resolution Steps:
Step 1: Declare P1 incident immediately
Step 2: Notify all city coordinators within 30 minutes
Step 3: Activate rail freight alternatives
Step 4: Coordinate with Ministry of Road Transport
Step 5: Monitor fuel/essential goods supply daily
Step 6: Escalate to National Head if beyond 48 hours

═══════════════════════════════════════════════════════
KEDB-003 | Auto Rickshaw Strike - South India
═══════════════════════════════════════════════════════
Error ID        : KEDB-003
Category        : Auto Strike
Affected Cities : Chennai, Bangalore, Hyderabad, Kochi
Trigger         : Meter revision, app-based competition, fuel costs
Frequency       : 3-4 times per year
Typical Duration: 1-2 days
Priority Level  : P3

Root Cause:
Auto unions protest against app-based aggregators
(Ola, Uber, Rapido) taking market share and
demand for revised meter rates.

Known Workarounds:
1. Promote app-based cab alternatives
2. Coordinate with metro/bus for route coverage
3. Issue advisory for early morning/late night travelers
4. Negotiate with union leaders via city transport dept

═══════════════════════════════════════════════════════
KEDB-004 | Bandh - Political
═══════════════════════════════════════════════════════
Error ID        : KEDB-004
Category        : Political Bandh
Affected Cities : Varies by state
Trigger         : Political events, court verdicts, protests
Frequency       : Unpredictable
Typical Duration: 1 day (usually)
Priority Level  : P1 (Bharat Bandh) / P2 (State Bandh)

Root Cause:
Political parties call bandh in response to
government decisions, court verdicts, or
national/state-level political events.

Known Workarounds:
1. Monitor political news 48 hours in advance
2. Pre-position essential service vehicles
3. Coordinate with police for essential movement
4. Activate work-from-home advisories for corporates
5. Ensure hospitals, airports exempt from bandh

Early Warning Indicators:
- Political party press conferences
- Social media announcements
- News reports 24-48 hours before

═══════════════════════════════════════════════════════
KEDB-005 | Railway Strike
═══════════════════════════════════════════════════════
Error ID        : KEDB-005
Category        : Railway Disruption
Affected Cities : All cities with rail connectivity
Trigger         : Employee demands, safety protests
Frequency       : Rare (1 in 2-3 years)
Typical Duration: 1-3 days
Priority Level  : P1

Root Cause:
Railway employee unions (AIRF, NFIR) strike over
pay revision, working conditions, privatization concerns.

Known Workarounds:
1. Activate bus transport on major routes
2. Coordinate with airlines for emergency capacity
3. Essential services exemption via Railway Board
4. State government road transport activation

═══════════════════════════════════════════════════════
KEDB-006 | Weather-Induced Transport Disruption
═══════════════════════════════════════════════════════
Error ID        : KEDB-006
Category        : Weather Disruption
Affected Cities : Coastal cities during monsoon
Trigger         : Heavy rain, cyclone, flood
Frequency       : Seasonal (June-September, cyclone season)
Typical Duration: Hours to days
Priority Level  : P1 (Cyclone) / P2 (Heavy Rain)

Root Cause:
Extreme weather events causing road flooding,
visibility issues, and unsafe driving conditions.

Known Workarounds:
1. Monitor IMD weather alerts continuously
2. Pre-activate flood response protocol
3. Coordinate with NDRF for rescue operations
4. Issue travel advisories 6 hours in advance
5. Activate alternate inland routes

Weather Thresholds for Action:
- Rain > 20mm/hr: Issue advisory
- Rain > 50mm/hr: Activate P2 protocol
- Cyclone warning: Activate P1 immediately
- Visibility < 50m: Halt operations
"""

    # ── 3. Resolution Logs ──────────────────────────────────
    resolution_content = """
PAST RESOLUTION LOGS
INDIA TRANSPORT MONITORING SYSTEM
Period: January 2023 - December 2023

════════════════════════════════════════════════════════
LOG-2023-001 | Mumbai Bus Strike
════════════════════════════════════════════════════════
Incident ID    : INC-2023-001
Date           : March 15, 2023
City           : Mumbai
Type           : Bus Strike (MSRTC)
Priority       : P2
KEDB Reference : KEDB-001

Timeline:
06:00 - Strike announced by MSRTC union
06:30 - P2 incident declared, team notified
07:00 - Backup transport MOU activated
07:30 - Public advisory issued on social media
08:00 - Mumbai Metro increased frequency by 40%
09:00 - App-based cabs deployed (2000+ vehicles)
12:00 - State transport minister initiated talks
18:00 - Partial agreement reached
20:00 - Buses resumed on 60% routes
Next Day 06:00 - Full service restored

Resolution Time: 24 hours
SLA Compliance: YES (P2 = 8hr resolution target, met with partial)
Root Cause: Wage revision demand (Rs 8000 increase)
Resolution: Rs 5000 interim increase, committee formed
Team: City Coordinator Rahul S, Manager Priya M

Lessons Learned:
- Early activation of metro backup was effective
- App-based cab coordination needs pre-signed agreements
- Social media advisory reached 2M users

════════════════════════════════════════════════════════
LOG-2023-002 | Delhi Truck Strike
════════════════════════════════════════════════════════
Incident ID    : INC-2023-002
Date           : June 1, 2023
City           : Delhi (National Impact)
Type           : National Truck Strike
Priority       : P1
KEDB Reference : KEDB-002

Timeline:
Day 1 00:00 - Strike begins (AIMTC announcement)
Day 1 00:30 - P1 declared, National Head notified
Day 1 01:00 - All city coordinators alerted
Day 1 02:00 - Rail freight alternatives activated
Day 1 06:00 - Essential goods supply monitored
Day 1 09:00 - Ministry of Transport meeting called
Day 2 - Fuel shortage reported in 8 cities
Day 2 - Army deployed for essential supplies
Day 3 - Government announced fuel price rollback
Day 3 18:00 - Strike called off
Day 4 - Normal operations resumed

Resolution Time: 72 hours
SLA Compliance: YES (P1 = 4hr response met, resolution extended)
Root Cause: Diesel price hike of Rs 12/litre
Resolution: Government rolled back Rs 8/litre, rest absorbed
Financial Impact: Rs 2.3 Crore estimated loss
Team: National Head Vikram R, Regional Directors x4

Lessons Learned:
- 48-hour advance warning was available but missed
- Rail freight activation saved critical supply chain
- Need better political intelligence monitoring

════════════════════════════════════════════════════════
LOG-2023-003 | Chennai Auto Strike
════════════════════════════════════════════════════════
Incident ID    : INC-2023-003
Date           : August 10, 2023
City           : Chennai
Type           : Auto Rickshaw Strike
Priority       : P3
KEDB Reference : KEDB-003

Timeline:
07:00 - Auto union announced strike
08:00 - P3 declared
08:30 - Advisory issued for app-based alternatives
09:00 - Metro and bus frequency increased
10:00 - Coordination with Ola/Uber for surge pricing cap
14:00 - Union leaders met transport commissioner
17:00 - Strike called off after meter revision promise

Resolution Time: 10 hours
SLA Compliance: YES (P3 = 24hr target, resolved in 10hr)
Root Cause: Demand for 30% meter rate revision
Resolution: 15% immediate revision, 15% after 6 months

════════════════════════════════════════════════════════
LOG-2023-004 | Bharat Bandh - Political
════════════════════════════════════════════════════════
Incident ID    : INC-2023-004
Date           : September 27, 2023
Cities         : All India
Type           : Bharat Bandh
Priority       : P1
KEDB Reference : KEDB-004

Timeline:
Day -2: Political party announced Bharat Bandh
Day -1: P1 pre-declared, all teams on standby
Day -1: WFH advisories sent to 500+ corporates
Day 0 00:00: Bandh begins
Day 0 06:00: Essential services confirmed operational
Day 0 09:00: Hospitals, airports functioning normally
Day 0 12:00: Partial relaxation in some states
Day 0 18:00: Bandh ends as scheduled
Day 0 20:00: Normal operations resumed

Resolution Time: 18 hours (as expected, single day)
SLA Compliance: YES
Pre-warning: 48 hours advance notice available
Team: National coordination team

Lessons Learned:
- 48-hour advance notice allows excellent preparation
- Pre-declared WFH was very effective
- Essential services exemption process worked well

════════════════════════════════════════════════════════
LOG-2023-005 | Mumbai Floods - Weather
════════════════════════════════════════════════════════
Incident ID    : INC-2023-005
Date           : July 20, 2023
City           : Mumbai
Type           : Weather-Induced Disruption
Priority       : P1
KEDB Reference : KEDB-006

Timeline:
06:00 - IMD issued red alert (200mm rain expected)
06:30 - P1 pre-declared based on weather alert
07:00 - Travel advisory issued
09:00 - Rain intensity: 45mm/hr (threshold exceeded)
10:00 - Local trains suspended
11:00 - Roads flooded in 12 areas
12:00 - NDRF deployed
15:00 - Rain reduced to 15mm/hr
17:00 - Local trains partially resumed
20:00 - Major roads cleared
Next Day 06:00 - Full normalcy

Resolution Time: 24 hours
SLA Compliance: YES (weather extension clause applied)
Rainfall: 187mm in 12 hours (record for July)
"""

    # ── 4. SOP Document ────────────────────────────────────
    sop_content = """
STANDARD OPERATING PROCEDURE
TRANSPORT STRIKE RESPONSE
Version 1.3 | India Transport Monitoring System

SOP-001: IMMEDIATE RESPONSE TO TRANSPORT STRIKE

TRIGGER: Strike news detected OR strike confirmed

STEP 1: VERIFY (0-15 minutes)
□ Check official union/government sources
□ Cross-verify with 2+ news sources
□ Confirm affected cities and routes
□ Assess number of vehicles/workers involved
□ Determine if political or economic cause

STEP 2: CLASSIFY PRIORITY (15-30 minutes)
□ Apply SLA priority matrix
□ P1: National/multi-metro, Bharat Bandh, Railway
□ P2: State/single metro, State Bandh
□ P3: City partial, Auto/Taxi strike
□ P4: Single route, minor disruption

STEP 3: NOTIFY STAKEHOLDERS (30-60 minutes)
□ Send incident notification email
□ Update incident management system
□ Notify city coordinators (P1/P2)
□ Notify National Head (P1 only)
□ Post public advisory on social media

STEP 4: ACTIVATE BACKUP TRANSPORT
□ Check KEDB for known workarounds
□ Activate pre-signed MOU with backup operators
□ Coordinate with metro/rail for increased frequency
□ Deploy app-based cab emergency agreements
□ Arrange government transport for essential services

STEP 5: MONITOR AND UPDATE
□ Check situation every 2 hours (P1/P2)
□ Check situation every 4 hours (P3/P4)
□ Update stakeholders at each check
□ Document all actions in resolution log
□ Track SLA compliance timers

STEP 6: ESCALATION TRIGGERS
□ No resolution after 50% of SLA time → escalate one level
□ Situation worsening → escalate immediately
□ Media attention increasing → notify PR team
□ Essential services affected → notify government

STEP 7: RESOLUTION AND CLOSURE
□ Confirm full service restoration
□ Document resolution steps taken
□ Calculate total resolution time
□ Check SLA compliance
□ Update KEDB with new learnings
□ Send closure notification to stakeholders
□ File incident report within 24 hours

SOP-002: WEATHER-INDUCED DISRUPTION RESPONSE

TRIGGER: Weather alert from IMD OR rain > 20mm/hr detected

STEP 1: MONITOR WEATHER
□ Check IMD alerts every 2 hours during monsoon
□ Set automatic alerts for rain > 15mm/hr
□ Monitor cyclone tracker during Oct-Dec

STEP 2: PRE-EMPTIVE ACTION (when warning issued)
□ Issue travel advisory 6 hours before
□ Notify transport operators
□ Pre-position emergency vehicles
□ Coordinate with NDRF if cyclone/flood risk

STEP 3: DURING EVENT
□ Real-time monitoring every 30 minutes
□ Update advisories as situation changes
□ Coordinate with local administration
□ Track road closures and flooding

STEP 4: RECOVERY
□ Assess damage to infrastructure
□ Coordinate road clearing operations
□ Restore transport in phases
□ Document weather data for future reference

CONTACT LIST:
National Head         : +91-XXXXXXXXXX
Regional Director N   : +91-XXXXXXXXXX
Regional Director S   : +91-XXXXXXXXXX
Regional Director E   : +91-XXXXXXXXXX
Regional Director W   : +91-XXXXXXXXXX
IMD Emergency         : 1800-180-1717
NDRF Control Room     : 011-24363260
"""

    # ── Save all documents ──────────────────────────────────
    docs = {
        "sla_2024.txt"           : sla_content,
        "kedb_2024.txt"          : kedb_content,
        "resolution_logs_2023.txt": resolution_content,
        "sop_strike_response.txt" : sop_content,
    }

    for filename, content in docs.items():
        path = DOCS_DIR / filename
        with open(path, "w", encoding="utf-8") as f:
            f.write(content.strip())
        print(f"  ✅ Created: {filename}")

    print(f"\n📂 All documents saved to: {DOCS_DIR}")
    print("   Replace these with your real documents anytime.")
    return list(docs.keys())


# ── Run ────────────────────────────────────────────────────
print("📄 Creating sample documents...")
created = create_sample_documents()
print(f"\n✅ {len(created)} sample documents created")
print("\n💡 To use your own documents:")
print(f"   Drop them in: {DOCS_DIR}")
print("   Supported: .txt .pdf .docx .csv .xlsx")

# COMMAND ----------

# ============================================================
# CELL 14 | DOCUMENT LOADER
# Reads all document types and extracts clean text
# ============================================================

class DocumentLoader:
    """
    Loads documents from the documents folder.
    Supports: TXT, PDF, DOCX, CSV, XLSX
    Returns clean text with metadata.
    """

    SUPPORTED = [".txt", ".pdf", ".docx", ".csv", ".xlsx"]

    def __init__(self, docs_dir: Path = DOCS_DIR):
        self.docs_dir = docs_dir

    # ── Public API ─────────────────────────────────────────

    def load_all(self) -> List[Dict]:
        """
        Load all documents from docs folder.
        Returns list of document dicts.
        """
        files = [
            f for f in self.docs_dir.iterdir()
            if f.suffix.lower() in self.SUPPORTED
        ]

        if not files:
            print(f"⚠️  No documents found in {self.docs_dir}")
            return []

        print(f"📂 Loading {len(files)} documents...")
        documents = []

        for file in files:
            print(f"   📄 {file.name}...", end=" ")
            doc = self._load_file(file)
            if doc:
                documents.append(doc)
                print(f"✅ {len(doc['content'])} chars")
            else:
                print("❌ failed")

        print(f"\n✅ Loaded {len(documents)} documents")
        return documents

    def load_one(self, filename: str) -> Optional[Dict]:
        """Load a single document by filename"""
        path = self.docs_dir / filename
        if not path.exists():
            print(f"❌ File not found: {filename}")
            return None
        return self._load_file(path)

    # ── Private loaders ────────────────────────────────────

    def _load_file(self, path: Path) -> Optional[Dict]:
        """Route to correct loader based on extension"""
        ext = path.suffix.lower()
        loaders = {
            ".txt"  : self._load_txt,
            ".pdf"  : self._load_pdf,
            ".docx" : self._load_docx,
            ".csv"  : self._load_csv,
            ".xlsx" : self._load_xlsx,
        }
        loader = loaders.get(ext)
        if not loader:
            return None

        try:
            content = loader(path)
            if not content or not content.strip():
                return None

            return {
                "filename"    : path.name,
                "filepath"    : str(path),
                "extension"   : ext,
                "category"    : self._detect_category(path.name),
                "content"     : content.strip(),
                "char_count"  : len(content),
                "loaded_at"   : datetime.now().isoformat(),
                "doc_hash"    : hashlib.md5(
                    content.encode()
                ).hexdigest()[:8],
            }
        except Exception as e:
            print(f"❌ Error loading {path.name}: {e}")
            return None

    def _load_txt(self, path: Path) -> str:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

    def _load_pdf(self, path: Path) -> str:
        try:
            reader = PdfReader(str(path))
            text   = []
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    text.append(t)
            return "\n".join(text)
        except Exception as e:
            print(f"PDF error: {e}")
            return ""

    def _load_docx(self, path: Path) -> str:
        try:
            doc   = docx.Document(str(path))
            paras = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paras)
        except Exception as e:
            print(f"DOCX error: {e}")
            return ""

    def _load_csv(self, path: Path) -> str:
        try:
            df   = pd.read_csv(path)
            rows = []
            for _, row in df.iterrows():
                row_text = " | ".join(
                    f"{col}: {val}"
                    for col, val in row.items()
                    if pd.notna(val)
                )
                rows.append(row_text)
            return "\n".join(rows)
        except Exception as e:
            print(f"CSV error: {e}")
            return ""

    def _load_xlsx(self, path: Path) -> str:
        try:
            xl   = pd.ExcelFile(path)
            all_text = []
            for sheet in xl.sheet_names:
                df   = xl.parse(sheet)
                all_text.append(f"[Sheet: {sheet}]")
                for _, row in df.iterrows():
                    row_text = " | ".join(
                        f"{col}: {val}"
                        for col, val in row.items()
                        if pd.notna(val)
                    )
                    all_text.append(row_text)
            return "\n".join(all_text)
        except Exception as e:
            print(f"XLSX error: {e}")
            return ""

    def _detect_category(self, filename: str) -> str:
        """Detect document category from filename"""
        fname = filename.lower()
        if "sla"        in fname: return "sla"
        if "kedb"       in fname: return "kedb"
        if "resolution" in fname: return "resolution_log"
        if "sop"        in fname: return "sop"
        if "policy"     in fname: return "policy"
        return "general"


# ── Run ────────────────────────────────────────────────────
loader    = DocumentLoader()
documents = loader.load_all()

print("\n📊 Document Summary:")
for doc in documents:
    print(f"   {doc['filename']:<40} "
          f"| {doc['category']:<15} "
          f"| {doc['char_count']:>6} chars")

# COMMAND ----------

# ============================================================
# CELL 15 | TEXT CHUNKER
# Splits documents into small overlapping pieces
# ============================================================

class TextChunker:
    """
    Splits long documents into smaller chunks.

    Why chunks?
        Embedding models have token limits.
        Smaller chunks = more precise retrieval.
        Overlap ensures context is not lost at boundaries.
    """

    def __init__(
        self,
        chunk_size    : int = CHUNK_SIZE,
        chunk_overlap : int = CHUNK_OVERLAP
    ):
        self.chunk_size    = chunk_size
        self.chunk_overlap = chunk_overlap

    # ── Public API ─────────────────────────────────────────

    def chunk_documents(self, documents: List[Dict]) -> List[Dict]:
        """
        Chunk all documents.
        Returns flat list of chunk dicts.
        """
        all_chunks = []
        for doc in documents:
            chunks = self._chunk_one(doc)
            all_chunks.extend(chunks)
            print(f"   📄 {doc['filename']:<40} → {len(chunks)} chunks")

        print(f"\n✅ Total chunks created: {len(all_chunks)}")
        return all_chunks

    # ── Private ────────────────────────────────────────────

    def _chunk_one(self, doc: Dict) -> List[Dict]:
        """Split one document into chunks with metadata"""
        text   = doc["content"]
        chunks = []

        # Split by paragraphs first (smarter than pure char split)
        paragraphs = self._split_paragraphs(text)
        current    = ""
        chunk_idx  = 0

        for para in paragraphs:
            # If adding this paragraph exceeds chunk size
            if len(current) + len(para) > self.chunk_size and current:
                # Save current chunk
                chunks.append(
                    self._make_chunk(doc, current, chunk_idx)
                )
                chunk_idx += 1

                # Start new chunk with overlap
                overlap_text = current[-self.chunk_overlap:]
                current      = overlap_text + "\n" + para
            else:
                current += "\n" + para if current else para

        # Save last chunk
        if current.strip():
            chunks.append(
                self._make_chunk(doc, current, chunk_idx)
            )

        return chunks

    def _split_paragraphs(self, text: str) -> List[str]:
        """Split text into paragraphs, filter empty ones"""
        paras = text.split("\n")
        return [p.strip() for p in paras if p.strip()]

    def _make_chunk(
        self, doc: Dict, text: str, idx: int
    ) -> Dict:
        """Create a chunk dict with full metadata"""
        return {
            # Content
            "chunk_id"      : f"{doc['doc_hash']}_{idx:03d}",
            "text"          : text.strip(),
            "char_count"    : len(text),

            # Source metadata (critical for citations)
            "source_file"   : doc["filename"],
            "source_path"   : doc["filepath"],
            "category"      : doc["category"],
            "chunk_index"   : idx,
            "doc_hash"      : doc["doc_hash"],

            # For display
            "preview"       : text[:100].replace("\n", " ") + "...",
        }

    def save_chunks(self, chunks: List[Dict]) -> str:
        """Save chunks to JSON for inspection"""
        path = CHUNKS_DIR / "all_chunks.json"
        with open(path, "w") as f:
            json.dump(chunks, f, indent=2)
        print(f"💾 Chunks saved to: {path}")
        return str(path)


# ── Run ────────────────────────────────────────────────────
print("✂️  Chunking documents...")
chunker    = TextChunker()
all_chunks = chunker.chunk_documents(documents)
chunker.save_chunks(all_chunks)

print(f"\n📊 Chunk Stats:")
categories = {}
for c in all_chunks:
    cat = c["category"]
    categories[cat] = categories.get(cat, 0) + 1
for cat, count in categories.items():
    print(f"   {cat:<20}: {count} chunks")

# COMMAND ----------

# ============================================================
# CELL 16 | EMBEDDING ENGINE + VECTOR STORE
# Converts chunks to vectors and stores in FAISS
# ============================================================

class VectorStore:
    """
    Converts text chunks to embeddings
    and stores them in FAISS for fast similarity search.

    Uses sentence-transformers locally.
    No API key, no internet after first download.
    """

    INDEX_FILE  = VECTOR_DIR / "index.faiss"
    META_FILE   = VECTOR_DIR / "metadata.pkl"
    CHUNKS_FILE = VECTOR_DIR / "chunks.pkl"

    def __init__(self, model_name: str = EMBEDDING_MODEL):
        print(f"🤖 Loading embedding model: {model_name}")
        print("   (Downloads ~90MB first time, then cached)")
        self.model  = SentenceTransformer(model_name)
        self.dim    = self.model.get_sentence_embedding_dimension()
        self.index  = None
        self.chunks = []
        print(f"✅ Model loaded | Embedding dimension: {self.dim}")

    # ── Public API ─────────────────────────────────────────

    def build(self, chunks: List[Dict]) -> None:
        """
        Build vector index from chunks.
        Call this when you add new documents.
        """
        print(f"\n🔨 Building vector index...")
        print(f"   Chunks to embed: {len(chunks)}")

        # Extract text for embedding
        texts = [c["text"] for c in chunks]

        # Create embeddings in batches
        print("   Creating embeddings (this takes 1-2 minutes)...")
        embeddings = self.model.encode(
            texts,
            batch_size    = 32,
            show_progress_bar = True,
            convert_to_numpy  = True,
        )

        # Normalize for cosine similarity
        faiss.normalize_L2(embeddings)

        # Build FAISS index
        self.index  = faiss.IndexFlatIP(self.dim)  # Inner product = cosine after normalize
        self.index.add(embeddings.astype(np.float32))
        self.chunks = chunks

        print(f"\n✅ Index built!")
        print(f"   Vectors stored : {self.index.ntotal}")
        print(f"   Embedding dim  : {self.dim}")

        # Save to disk
        self._save()

    def search(
        self,
        query   : str,
        top_k   : int = TOP_K_RESULTS,
        category: str = None
    ) -> List[Dict]:
        """
        Search for most relevant chunks.

        Parameters:
            query    → your question in plain English
            top_k    → how many results to return
            category → filter by doc type (sla/kedb/resolution_log/sop)

        Returns list of chunks with similarity scores.
        """
        if self.index is None:
            print("⚠️  Index not built. Run build() first.")
            return []

        # Embed the query
        q_vec = self.model.encode(
            [query],
            convert_to_numpy=True
        ).astype(np.float32)
        faiss.normalize_L2(q_vec)

        # Search — get more results if filtering by category
        search_k = top_k * 4 if category else top_k
        scores, indices = self.index.search(q_vec, search_k)

        # Build results
        results = []
        for score, idx in zip(scores[0], indices[0]):
            if idx == -1:
                continue
            chunk = self.chunks[idx].copy()
            chunk["similarity_score"] = float(score)

            # Apply category filter
            if category and chunk["category"] != category:
                continue

            results.append(chunk)
            if len(results) >= top_k:
                break

        return results

    def load(self) -> bool:
        """Load existing index from disk"""
        if not self.INDEX_FILE.exists():
            return False
        try:
            self.index  = faiss.read_index(str(self.INDEX_FILE))
            with open(self.META_FILE,  "rb") as f:
                meta = pickle.load(f)
            with open(self.CHUNKS_FILE, "rb") as f:
                self.chunks = pickle.load(f)
            print(f"✅ Index loaded from disk")
            print(f"   Vectors: {self.index.ntotal}")
            return True
        except Exception as e:
            print(f"❌ Load failed: {e}")
            return False

    def add_documents(self, new_chunks: List[Dict]) -> None:
        """Add new chunks to existing index without rebuilding"""
        if not new_chunks:
            return

        texts      = [c["text"] for c in new_chunks]
        embeddings = self.model.encode(
            texts,
            convert_to_numpy=True
        ).astype(np.float32)
        faiss.normalize_L2(embeddings)

        self.index.add(embeddings)
        self.chunks.extend(new_chunks)
        self._save()
        print(f"✅ Added {len(new_chunks)} new chunks to index")

    # ── Private ────────────────────────────────────────────

    def _save(self) -> None:
        """Save index and metadata to disk"""
        faiss.write_index(self.index, str(self.INDEX_FILE))
        with open(self.META_FILE, "wb") as f:
            pickle.dump({"total": self.index.ntotal}, f)
        with open(self.CHUNKS_FILE, "wb") as f:
            pickle.dump(self.chunks, f)
        print(f"💾 Index saved to: {VECTOR_DIR}")


# ── Run ────────────────────────────────────────────────────
vector_store = VectorStore()

# Try loading existing index first
if not vector_store.load():
    # Build fresh if not found
    vector_store.build(all_chunks)

print("\n✅ Cell 16 complete — Vector store ready")

# COMMAND ----------

# ============================================================
# CELL 17 | RAG QUERY ENGINE
# Main interface for asking questions
# ============================================================

class RAGQueryEngine:
    """
    Main RAG interface.
    Ask questions in plain English.
    Get answers with source citations.
    """

    def __init__(self, vector_store: VectorStore):
        self.vs = vector_store

    # ── Public API ─────────────────────────────────────────

    def query(
        self,
        question : str,
        category : str = None,
        top_k    : int = TOP_K_RESULTS,
        verbose  : bool = True
    ) -> Dict:
        """
        Ask any question about your documents.

        Parameters:
            question → plain English question
            category → "sla" / "kedb" / "resolution_log" / "sop"
            top_k    → number of source chunks to use
            verbose  → print results to screen

        Returns:
            {
                question  : str,
                answer    : str,
                sources   : list,
                category  : str,
                chunks    : list (raw chunks)
            }
        """
        # Retrieve relevant chunks
        chunks = self.vs.search(question, top_k=top_k, category=category)

        if not chunks:
            return self._empty_result(question)

        # Build answer from chunks
        answer  = self._synthesize(question, chunks)
        sources = self._format_sources(chunks)

        result = {
            "question"  : question,
            "answer"    : answer,
            "sources"   : sources,
            "category"  : category or "all",
            "chunks"    : chunks,
            "chunk_count": len(chunks),
        }

        if verbose:
            self._print_result(result)

        return result

    def query_sla(self, question: str) -> Dict:
        """Query only SLA documents"""
        return self.query(question, category="sla")

    def query_kedb(self, question: str) -> Dict:
        """Query only KEDB documents"""
        return self.query(question, category="kedb")

    def query_resolution(self, question: str) -> Dict:
        """Query only resolution logs"""
        return self.query(question, category="resolution_log")

    def query_sop(self, question: str) -> Dict:
        """Query only SOP documents"""
        return self.query(question, category="sop")

    # ── Private ────────────────────────────────────────────

    def _synthesize(self, question: str, chunks: List[Dict]) -> str:
        """
        Build answer from retrieved chunks.
        (Rule-based synthesis — no LLM needed)
        """
        # Combine all chunk texts
        context = "\n\n---\n\n".join(
            f"[{c['category'].upper()} | {c['source_file']}]\n{c['text']}"
            for c in chunks
        )

        # Find most relevant sentences
        q_words    = set(question.lower().split())
        sentences  = []

        for chunk in chunks:
            for sent in chunk["text"].split("."):
                sent = sent.strip()
                if len(sent) < 20:
                    continue
                # Score sentence by word overlap with question
                s_words = set(sent.lower().split())
                overlap = len(q_words & s_words)
                if overlap > 0:
                    sentences.append((overlap, sent, chunk))

        # Sort by relevance
        sentences.sort(key=lambda x: x[0], reverse=True)

        if not sentences:
            return chunks[0]["text"][:500]

        # Build answer from top sentences
        seen   = set()
        answer = []
        for _, sent, chunk in sentences[:5]:
            key = sent[:50]
            if key not in seen:
                seen.add(key)
                answer.append(sent)

        return ". ".join(answer) + "."

    def _format_sources(self, chunks: List[Dict]) -> List[Dict]:
        """Format source citations"""
        seen    = set()
        sources = []
        for c in chunks:
            key = c["source_file"]
            if key not in seen:
                seen.add(key)
                sources.append({
                    "file"      : c["source_file"],
                    "category"  : c["category"],
                    "relevance" : f"{c['similarity_score']:.2%}",
                    "preview"   : c["preview"],
                })
        return sources

    def _empty_result(self, question: str) -> Dict:
        return {
            "question"   : question,
            "answer"     : "No relevant information found in documents.",
            "sources"    : [],
            "category"   : "none",
            "chunks"     : [],
            "chunk_count": 0,
        }

    def _print_result(self, result: Dict) -> None:
        print(f"\n{'='*55}")
        print(f"❓ Question : {result['question']}")
        print(f"{'='*55}")
        print(f"\n💡 Answer:")
        print(f"   {result['answer']}")
        print(f"\n📚 Sources ({len(result['sources'])}):")
        for s in result["sources"]:
            print(f"   📄 {s['file']} | {s['category']} | relevance: {s['relevance']}")
        print(f"{'='*55}")


# ── Run ────────────────────────────────────────────────────
rag = RAGQueryEngine(vector_store)

# Test queries
print("🧪 Testing RAG queries...\n")
rag.query("What is the SLA response time for P1 transport strike?")
rag.query("What workaround was used for Mumbai bus strike?")
rag.query("What are the steps when a bandh is declared?")

# COMMAND ----------

# ============================================================
# CELL 18 | STRUCTURED OUTPUT FORMATTER
# Combines SQLite data + RAG results into ML-ready format
# ============================================================

class StructuredOutputBuilder:
    """
    Combines:
        - Strike data from SQLite (DatabaseManager)
        - Weather data from SQLite
        - SLA rules from RAG
        - KEDB matches from RAG
        - Resolution history from RAG

    Outputs clean structured dict ready for ML model.
    """

    def __init__(self, db: DatabaseManager, rag: RAGQueryEngine):
        self.db  = db
        self.rag = rag

    # ── Public API ─────────────────────────────────────────

    def build(
        self,
        strike_title : str,
        city         : str,
        date         : str,
        severity     : str = None,
        strike_type  : str = None,
    ) -> Dict:
        """
        Build complete ML-ready record for one strike incident.

        Parameters:
            strike_title → headline of the strike news
            city         → affected city
            date         → date of incident YYYY-MM-DD
            severity     → HIGH/MEDIUM/LOW (optional)
            strike_type  → bus/truck/railway etc (optional)

        Returns structured dict ready for ML.
        """

        print(f"\n🔨 Building structured output...")
        print(f"   Strike : {strike_title[:60]}...")
        print(f"   City   : {city} | Date: {date}")

        # ── 1. Weather context ─────────────────────────────
        weather = self._get_weather_context(city, date)

        # ── 2. SLA lookup ──────────────────────────────────
        sla = self._get_sla_context(strike_title, city, severity)

        # ── 3. KEDB lookup ─────────────────────────────────
        kedb = self._get_kedb_context(strike_title, strike_type)

        # ── 4. Resolution history ──────────────────────────
        history = self._get_resolution_context(city, strike_type)

        # ── 5. SOP lookup ──────────────────────────────────
        sop = self._get_sop_context(strike_title)

        # ── 6. Combine into ML record ──────────────────────
        record = self._combine(
            strike_title, city, date,
            severity, strike_type,
            weather, sla, kedb, history, sop
        )

        return record

    def build_batch(self, strikes_df: pd.DataFrame) -> pd.DataFrame:
        """
        Build structured output for all strikes in a DataFrame.
        Returns ML-ready DataFrame.
        """
        records = []
        total   = len(strikes_df)

        print(f"\n📦 Building batch output for {total} strikes...")

        for i, (_, row) in enumerate(strikes_df.iterrows()):
            print(f"   [{i+1}/{total}] {row.get('city_mentioned','?')}...", end=" ")
            try:
                record = self.build(
                    strike_title = row.get("title", ""),
                    city         = row.get("city_mentioned", "Unknown"),
                    date         = row.get("published_date", ""),
                    severity     = row.get("severity", "LOW"),
                    strike_type  = row.get("strike_type", "general"),
                )
                records.append(record)
                print("✅")
            except Exception as e:
                print(f"❌ {e}")

        df = pd.DataFrame(records)
        print(f"\n✅ Built {len(df)} structured records")
        return df

    def save(self, df: pd.DataFrame, filename: str = None) -> str:
        """Save ML-ready output to CSV and JSON"""
        if filename is None:
            filename = f"ml_ready_{datetime.now():%Y%m%d_%H%M%S}"

        # CSV
        csv_path = CSV_DIR / f"{filename}.csv"
        df.to_csv(csv_path, index=False)

        # JSON
        json_path = JSON_DIR / f"{filename}.json"
        payload   = {
            "generated_at" : datetime.now().isoformat(),
            "total_records": len(df),
            "columns"      : list(df.columns),
            "data"         : df.to_dict(orient="records"),
        }
        with open(json_path, "w") as f:
            json.dump(payload, f, indent=2, default=str)

        print(f"\n💾 ML-ready data saved:")
        print(f"   CSV  : {csv_path}")
        print(f"   JSON : {json_path}")
        return str(csv_path)

    # ── Private helpers ────────────────────────────────────

    def _get_weather_context(self, city: str, date: str) -> Dict:
        """Get weather data from SQLite for this city/date"""
        try:
            df = self.db.read_weather(date, city)
            if df.empty:
                return self._empty_weather()
            return {
                "avg_temp"     : round(df["temperature"].mean(), 1),
                "max_temp"     : round(df["temperature"].max(), 1),
                "min_temp"     : round(df["temperature"].min(), 1),
                "avg_humidity" : round(df["humidity"].mean(), 1),
                "total_rain"   : round(df["rain_1h"].sum(), 2),
                "condition"    : df["weather_description"].mode()[0]
                                 if not df.empty else "unknown",
                "has_rain"     : int(df["rain_1h"].sum() > 0),
                "has_storm"    : int(
                    df["weather_main"].str.lower()
                    .str.contains("storm").any()
                ),
            }
        except:
            return self._empty_weather()

    def _get_sla_context(
        self, title: str, city: str, severity: str
    ) -> Dict:
        """Get SLA rules from RAG"""
        query  = f"SLA response time resolution time for {severity or ''} transport strike {city}"
        result = self.rag.query_sla(query)

        priority = self._map_severity_to_priority(severity)

        return {
            "sla_priority"       : priority,
            "sla_relevant_text"  : result["answer"][:300],
            "sla_source"         : result["sources"][0]["file"]
                                   if result["sources"] else "unknown",
            "sla_found"          : int(result["chunk_count"] > 0),
        }

    def _get_kedb_context(
        self, title: str, strike_type: str
    ) -> Dict:
        """Get KEDB workarounds from RAG"""
        query  = f"known error workaround solution for {strike_type or ''} strike"
        result = self.rag.query_kedb(query)

        # Extract KEDB ID if mentioned
        kedb_id = "unknown"
        for chunk in result.get("chunks", []):
            import re
            match = re.search(r"KEDB-\d+", chunk["text"])
            if match:
                kedb_id = match.group()
                break

        return {
            "kedb_id"           : kedb_id,
            "kedb_match_found"  : int(result["chunk_count"] > 0),
            "kedb_workaround"   : result["answer"][:300],
            "kedb_source"       : result["sources"][0]["file"]
                                  if result["sources"] else "unknown",
        }

    def _get_resolution_context(
        self, city: str, strike_type: str
    ) -> Dict:
        """Get past resolution data from RAG"""
        query  = f"past resolution {city} {strike_type or ''} strike how was it resolved"
        result = self.rag.query_resolution(query)

        # Try to extract resolution time from text
        import re
        res_time = None
        for chunk in result.get("chunks", []):
            match = re.search(r"Resolution Time:\s*(\d+)\s*hours?", chunk["text"])
            if match:
                res_time = int(match.group(1))
                break

        return {
            "past_resolution_found"   : int(result["chunk_count"] > 0),
            "past_resolution_summary" : result["answer"][:300],
            "past_resolution_time_hr" : res_time,
            "resolution_source"       : result["sources"][0]["file"]
                                        if result["sources"] else "unknown",
        }

    def _get_sop_context(self, title: str) -> Dict:
        """Get SOP steps from RAG"""
        query  = f"steps procedure response to transport strike"
        result = self.rag.query_sop(query)

        return {
            "sop_found"   : int(result["chunk_count"] > 0),
            "sop_steps"   : result["answer"][:300],
            "sop_source"  : result["sources"][0]["file"]
                            if result["sources"] else "unknown",
        }

    def _combine(
        self, title, city, date,
        severity, strike_type,
        weather, sla, kedb, history, sop
    ) -> Dict:
        """Combine all context into final ML record"""
        return {
            # ── Identifiers ───────────────────────────────
            "incident_title"          : title[:200],
            "city"                    : city,
            "date"                    : date,

            # ── Strike Classification ─────────────────────
            "severity"                : severity or "UNKNOWN",
            "strike_type"             : strike_type or "general",
            "is_high_severity"        : int(severity == "HIGH"),
            "is_national"             : int(city == "National/Unknown"),

            # ── Weather Features ──────────────────────────
            "weather_avg_temp"        : weather["avg_temp"],
            "weather_max_temp"        : weather["max_temp"],
            "weather_humidity"        : weather["avg_humidity"],
            "weather_total_rain"      : weather["total_rain"],
            "weather_has_rain"        : weather["has_rain"],
            "weather_has_storm"       : weather["has_storm"],
            "weather_condition"       : weather["condition"],

            # ── SLA Features ──────────────────────────────
            "sla_priority"            : sla["sla_priority"],
            "sla_found"               : sla["sla_found"],
            "sla_relevant_text"       : sla["sla_relevant_text"],

            # ── KEDB Features ─────────────────────────────
            "kedb_id"                 : kedb["kedb_id"],
            "kedb_match_found"        : kedb["kedb_match_found"],
            "kedb_workaround"         : kedb["kedb_workaround"],

            # ── Resolution History Features ───────────────
            "past_resolution_found"   : history["past_resolution_found"],
            "past_resolution_time_hr" : history["past_resolution_time_hr"],
            "past_resolution_summary" : history["past_resolution_summary"],

            # ── SOP Features ──────────────────────────────
            "sop_found"               : sop["sop_found"],
            "sop_steps"               : sop["sop_steps"],

            # ── Metadata ──────────────────────────────────
            "processed_at"            : datetime.now().isoformat(),
        }

    def _map_severity_to_priority(self, severity: str) -> str:
        return {
            "HIGH"   : "P1",
            "MEDIUM" : "P2",
            "LOW"    : "P3",
        }.get(severity, "P4")

    def _empty_weather(self) -> Dict:
        return {
            "avg_temp": None, "max_temp": None,
            "min_temp": None, "avg_humidity": None,
            "total_rain": 0,  "condition": "unknown",
            "has_rain": 0,    "has_storm": 0,
        }


# ── Run ────────────────────────────────────────────────────
output_builder = StructuredOutputBuilder(db, rag)

# Test with one strike
test_record = output_builder.build(
    strike_title = "MSRTC bus strike disrupts Mumbai transport",
    city         = "Mumbai",
    date         = datetime.now().strftime("%Y-%m-%d"),
    severity     = "HIGH",
    strike_type  = "bus",
)

print("\n📊 Sample ML-Ready Record:")
for key, val in test_record.items():
    if val and str(val)[:1] != "{":
        print(f"  {key:<30}: {str(val)[:60]}")

print("\n✅ Cell 18 complete!")

# COMMAND ----------

# ============================================================
# CELL 19 | FULL PIPELINE INTEGRATION
# Runs everything end to end
# ============================================================

def run_rag_pipeline(date: str = None, city: str = None):
    """
    Complete pipeline:
    1. Get strikes from DB (or API if missing)
    2. Run RAG on each strike
    3. Build ML-ready structured output
    4. Save everything
    """
    date = date or datetime.now().strftime("%Y-%m-%d")

    print("=" * 55)
    print("🚀 RAG PIPELINE")
    print(f"   Date: {date} | City: {city or 'All'}")
    print("=" * 55)

    # ── Step 1: Get strikes ────────────────────────────────
    print("\n[1/4] Fetching strike data...")
    result     = fetcher.get(date, city)
    strikes_df = result["strikes"]

    if strikes_df.empty:
        print("   ⚠️  No strikes found for this date")
        return None

    print(f"   ✅ {len(strikes_df)} strikes found")

    # ── Step 2: Run RAG on each strike ─────────────────────
    print("\n[2/4] Running RAG analysis...")
    ml_df = output_builder.build_batch(strikes_df)

    # ── Step 3: Save ML-ready output ──────────────────────
    print("\n[3/4] Saving ML-ready output...")
    tag      = date.replace("-", "")
    csv_path = output_builder.save(
        ml_df,
        filename=f"ml_ready_{tag}"
    )

    # ── Step 4: Summary ───────────────────────────────────
    print("\n[4/4] Summary:")
    print(f"   Total records    : {len(ml_df)}")
    print(f"   With KEDB match  : {ml_df['kedb_match_found'].sum()}")
    print(f"   With SLA info    : {ml_df['sla_found'].sum()}")
    print(f"   With past history: {ml_df['past_resolution_found'].sum()}")
    print(f"   High severity    : {ml_df['is_high_severity'].sum()}")

    print("\n" + "=" * 55)
    print("✅ RAG PIPELINE COMPLETE")
    print(f"   ML-ready file: {csv_path}")
    print("=" * 55)

    return ml_df


# ── Run ────────────────────────────────────────────────────
ml_ready_df = run_rag_pipeline()

if ml_ready_df is not None:
    print("\n📊 ML-Ready DataFrame Preview:")
    print(ml_ready_df[[
        "city", "severity", "strike_type",
        "sla_priority", "kedb_id",
        "past_resolution_time_hr",
        "weather_avg_temp"
    ]].to_string(index=False))

# COMMAND ----------

# ============================================================
# CELL 20 | ADD NEW DOCUMENTS ANYTIME
# Run this whenever you add new documents to the folder
# ============================================================

def refresh_rag():
    """
    Reload all documents and rebuild vector index.
    Run this when you:
    - Add new documents
    - Update existing documents
    - Add new KEDB entries
    """
    print("🔄 Refreshing RAG index...")
    print("=" * 45)

    # Reload documents
    print("\n[1/3] Loading documents...")
    docs   = loader.load_all()

    # Rechunk
    print("\n[2/3] Chunking...")
    chunks = chunker.chunk_documents(docs)
    chunker.save_chunks(chunks)

    # Rebuild index
    print("\n[3/3] Rebuilding vector index...")
    vector_store.build(chunks)

    print("\n✅ RAG refresh complete!")
    print(f"   Documents : {len(docs)}")
    print(f"   Chunks    : {len(chunks)}")
    print(f"   Vectors   : {vector_store.index.ntotal}")


# ── Quick query interface ───────────────────────────────────
def ask(question: str, category: str = None):
    """
    Quick way to ask any question.

    Examples:
        ask("What is P1 response time?")
        ask("What workaround for truck strike?", category="kedb")
        ask("How was Mumbai strike resolved?", category="resolution_log")
    """
    return rag.query(question, category=category)


print("✅ Cell 20 ready!")
print("\n💡 Quick commands:")
print('   ask("your question here")')
print('   ask("your question", category="sla")')
print('   ask("your question", category="kedb")')
print('   ask("your question", category="resolution_log")')
print('   ask("your question", category="sop")')
print('   refresh_rag()  ← when you add new documents')
print('   run_rag_pipeline("2024-01-15")  ← full pipeline for any date')